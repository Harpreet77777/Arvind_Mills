from datetime import date
from datetime import datetime
import calendar
import openpyxl
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from . import crud, models
import pandas as pd
from fastapi import HTTPException
from .routers import dyeing_report, preparatory_report
from openpyxl.styles import Border, Side, Alignment, PatternFill
import os
import sys
import traceback
from .routers.preparatory_report import get_production_kg_of_each_machine
from collections import defaultdict
from openpyxl.styles import Font, Alignment

if getattr(sys, 'frozen', False):
    dirname = os.path.dirname(sys.executable)
else:
    dirname = os.path.dirname(os.path.abspath(__file__))


def get_last_filled_row(ws):
    for row in range(ws.max_row, 0, -1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is not None:
                return row
    return 0


def copy_row(ws, src_row, dest_row):
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dest_cell = ws.cell(row=dest_row, column=col)

        dest_cell.value = src_cell.value
        dest_cell.font = src_cell.font
        dest_cell.border = src_cell.border
        dest_cell.fill = src_cell.fill
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = src_cell.protection
        dest_cell.alignment = src_cell.alignment


async def get_finishing_report_data(db: Session, date_: date):
    month_name = date_.strftime("%B")
    year = date_.year
    print("month_name==>", month_name)

    wb = load_workbook(os.path.join(dirname, f"Reports/Template/Finishing Section Report-IIoT.xlsx"))
    print("wb path :->", wb)
    ws2 = wb["Monthly Prod. Report"]
    ws2['A3'].value = month_name

    ws3 = wb["Airo Prod (AT 1 & 2)"]
    ws3['A2'].value = month_name

    ws4 = wb["heat consumption"]
    ws4['A3'].value = month_name

    ws6 = wb["Reprocess & demand report"]
    ws7 = wb["Airo (Effective & Non-Effective"]

    po_data = await crud.get_po_data_speed(db, date_)
    ws8 = wb["Heat Set Speed & Temp"]

    thin_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Write data to the sheet
    for i, datap in enumerate(po_data):
        try:
            row = i + 2

            ws8[f'A{row}'] = datap['date_']
            ws8[f'B{row}'] = datap['machine']
            ws8[f'C{row}'] = datap['k1']
            ws8[f'D{row}'] = datap['Finish']
            ws8[f'E{row}'] = datap['k4']
            ws8[f'F{row}'] = datap['po_number']
            ws8[f'G{row}'] = datap['po_qty']
            ws8[f'H{row}'] = datap['meter']
            ws8[f'J{row}'] = datap['finish_glm']
            ws8[f'K{row}'] = datap['temperature']
            ws8[f'L{row}'] = datap['width']
            ws8[f'M{row}'] = datap['speed']

        except Exception as e:
            print(e)

    # Apply border and alignment to the entire range
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M']:
        for row in range(2, len(po_data) + 2):
            cell = ws8[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = center_alignment

    ############################# MONTHLY PRODUCTION  REPORT DATA

    monthly_production = await crud.get_monthly_production(db, date_)

    column_mappings = {
        'Stenter-1': 2,
        'Stenter-2': 3,
        'Stenter-3': 4,
        'Stenter-4': 5,
        'Stenter-5': 6,
        'Stenter-6': 7,
        'Sanforiser-2': 9,
        'Sanforiser-3': 10,
        'Sanforiser-4': 11,
        'Relax-Drier': 13,
        'Airo-24': 14,
        'Curing': 15
    }

    ws2["A1"].value = f"Finishing Deptt.Production data for the month of {month_name}-{year}"
    for machine, machine_data in monthly_production.items():
        c_idx = column_mappings.get(machine)

        if c_idx is not None:
            row_idx = 5
            for data_dict in machine_data:
                date_str = data_dict['date']
                formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
                ws2.cell(row=row_idx, column=1, value=formatted_date)
                ws2.cell(row=row_idx, column=c_idx, value=data_dict['production'])
                row_idx += 1

    airo_production = await crud.get_airo_production(db, date_)
    for r_idx, entry in enumerate(airo_production, start=3):
        date_str = entry['date']
        date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
        machine = entry['machine']
        non_effective_production = entry['non_effective_production']
        production_total = entry['production_total']

        ws3[f'A{r_idx}'] = date
        ws3[f'B{r_idx}'] = int(non_effective_production)
        ws3[f'E{r_idx}'] = int(production_total)

    ############################# HEAT CONSUMPTION REPORT DATA
    heat_production = await crud.get_heat_consumption(db, date_)
    ws4.cell(row=1, column=1, value=f" Daily Heat Consumption Data {month_name}-{year}")
    for r_idx, entry in enumerate(heat_production, start=4):
        print(entry)
        date_str = entry['date']
        date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
        stenter_data_production = entry['stenter_data_production']
        heat_thermic_fluid = entry['heat_thermic_fluid']
        total_heat_thermic = entry['total_heat_thermic']
        thermofix_production = entry['thermofix_production']
        avg_glm = entry['avg_glm']

        ws4[f'A{r_idx}'] = date
        ws4[f'B{r_idx}'] = stenter_data_production
        ws4[f'C{r_idx}'] = avg_glm
        ws4[f'E{r_idx}'] = total_heat_thermic
        ws4[f'J{r_idx}'] = thermofix_production
        ws4[f'K{r_idx}'] = heat_thermic_fluid

    #    ############################## REPROCESS AND PROCESS REPORT DATA
    print("Production Start  for Reprocess")
    reprocess_and_process_data = await crud.get_reprocess_and_process_demand(db, date_)

    ws6.cell(row=1, column=1, value=f" STENTER RE - PROCESS & PROCESS DEMAND REPORT {month_name}-{year}")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')

    start_row = 16
    total_rows = ws6.max_row
    num_rows_to_delete = total_rows - start_row + 1

    merged_ranges_to_unmerge = []
    for merged_cell in ws6.merged_cells.ranges:
        min_row, max_row = merged_cell.min_row, merged_cell.max_row
        if min_row >= start_row:
            merged_ranges_to_unmerge.append(merged_cell)

    for merged_cell in merged_ranges_to_unmerge:
        ws6.unmerge_cells(str(merged_cell))

    for _ in range(num_rows_to_delete):
        ws6.delete_rows(start_row)

    #########################3 FILL REPROCESS  DATA
    last_filled_row = get_last_filled_row(ws6)
    print("Production Start  for Reprocess 1")

    ws6.cell(row=last_filled_row + 3, column=10, value="TOTAL PRODUCTION").alignment = alignment
    total_stenter_production_ftd = reprocess_and_process_data['total_stenter_production_ftd']
    ws6.cell(row=last_filled_row + 3, column=11, value=total_stenter_production_ftd).alignment = alignment
    total_stenter_production_mtd = reprocess_and_process_data['total_stenter_production_mtd']
    ws6.cell(row=last_filled_row + 3, column=12, value=total_stenter_production_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 4, column=10, value="REPROCESS PRODUCTION").alignment = alignment
    total_stenter_reprocess_ftd = reprocess_and_process_data['total_stenter_reprocess_ftd']
    ws6.cell(row=last_filled_row + 4, column=11, value=total_stenter_reprocess_ftd).alignment = alignment
    total_stenter_reprocess_mtd = reprocess_and_process_data['total_stenter_reprocess_mtd']
    ws6.cell(row=last_filled_row + 4, column=12, value=total_stenter_reprocess_mtd).alignment = alignment

    print("Production Start  for Reprocess 2")
    ws6.cell(row=last_filled_row + 5, column=10, value="PERCENTAGE").alignment = alignment
    # ws6.cell(row=last_filled_row + 5, column=11,value=(total_stenter_reprocess_ftd / total_stenter_production_ftd) * 100).alignment = alignmen
    # ws6.cell(row=last_filled_row + 5, column=12, value=(total_stenter_reprocess_mtd / total_stenter_production_mtd) * 100).alignment = alignment
    if total_stenter_production_ftd != 0:
        ws6.cell(row=last_filled_row + 5, column=11,
                 value=(total_stenter_reprocess_ftd / total_stenter_production_ftd) * 100).alignment = alignment
    else:
        ws6.cell(row=last_filled_row + 5, column=11, value="0").alignment = alignment  # Handle zero division case

    # Check for zero for the MTD calculation as well
    if total_stenter_production_mtd != 0:
        ws6.cell(row=last_filled_row + 5, column=12,
                 value=(total_stenter_reprocess_mtd / total_stenter_production_mtd) * 100).alignment = alignment
    else:
        ws6.cell(row=last_filled_row + 5, column=12, value="0").alignment = alignment  # Handle zero division case

    print("Production Start  for Reprocess 3")
    reprocess_data = reprocess_and_process_data['reprocess_data']
    for row_index, data_dict in enumerate(reprocess_data):
        current_row = start_row + row_index
        # Write date in column A
        ws6.cell(row=current_row, column=1, value=data_dict['date_'])
        ws6.cell(row=current_row, column=2, value=data_dict['s_no'])
        ws6.cell(row=current_row, column=3, value=data_dict['article'])
        ws6.cell(row=current_row, column=4, value=data_dict['po_number'])
        ws6.cell(row=current_row, column=5, value=data_dict['k3'])
        ws6.cell(row=current_row, column=6, value=data_dict['meter_production'])

        for col in range(1, 13):
            cell = ws6.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment

    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 1, column=5, value="Total").alignment = alignment
    ws6.cell(row=last_filled_row + 1, column=6, value=f"=SUM(F{start_row}:F{last_filled_row})")
    calculated_value_reprocess_data = ws6.cell(row=last_filled_row + 1, column=6).value
    print("calculated_value_reprocess_data", calculated_value_reprocess_data)

    for col in range(1, 13):
        cell = ws6.cell(row=last_filled_row + 1, column=col)
        cell.border = thin_border
        cell.alignment = alignment

    #########################3 FILL PROCESS DEMAND DATA
    last_filled_row = get_last_filled_row(ws6)
    ws6.merge_cells(start_row=last_filled_row + 2, start_column=1, end_row=last_filled_row + 2, end_column=8)
    ws6.cell(row=last_filled_row + 2, column=1, value="Day wise Piece Dyed Process Demand  ROUTE FABRIC")
    for col in range(1, 12):  # Columns A to H (1 to 8)
        cell = ws6.cell(row=last_filled_row + 2, column=col)
        cell.border = thin_border
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='FF5cd65c', end_color='FF5cd65c', fill_type='solid')

    last_filled_row = get_last_filled_row(ws6)

    headers = ["Date", "S.No", "Article", "PO Number", "K3", "Meter Production", "Reason", "Deptt."]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws6.cell(row=last_filled_row + 1, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='FF5cd65c', end_color='FF5cd65c', fill_type='solid')

    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 3, column=10, value="TOTAL PRODUCTION").alignment = alignment
    total_stenter_production_ftd = reprocess_and_process_data['total_stenter_production_ftd']
    ws6.cell(row=last_filled_row + 3, column=11, value=total_stenter_production_ftd).alignment = alignment
    total_stenter_production_mtd = reprocess_and_process_data['total_stenter_production_mtd']
    ws6.cell(row=last_filled_row + 3, column=12, value=total_stenter_production_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 4, column=10, value="PROCESS DEMAND PRODUCTION")
    total_stenter_process_demand_ftd = reprocess_and_process_data['total_stenter_process_demand_ftd']
    ws6.cell(row=last_filled_row + 4, column=11, value=total_stenter_process_demand_ftd).alignment = alignment
    total_stenter_process_demand_mtd = reprocess_and_process_data['total_stenter_process_demand_mtd']
    ws6.cell(row=last_filled_row + 4, column=12, value=total_stenter_process_demand_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 5, column=10, value="PERCENTAGE")
    # ws6.cell(row=last_filled_row + 5, column=11,value=(total_stenter_reprocess_ftd / total_stenter_process_demand_ftd) * 100).alignment = alignment
    # ws6.cell(row=last_filled_row + 5, column=12, value=(total_stenter_reprocess_mtd / total_stenter_process_demand_mtd) * 100).alignment = alignment
    if total_stenter_process_demand_ftd != 0:
        ws6.cell(row=last_filled_row + 5, column=11,
                 value=(total_stenter_reprocess_ftd / total_stenter_process_demand_ftd) * 100).alignment = alignment
    else:
        ws6.cell(row=last_filled_row + 5, column=11, value="0").alignment = alignment  # Handle zero division case

    # Check for zero for the MTD calculation as well
    if total_stenter_process_demand_mtd != 0:
        ws6.cell(row=last_filled_row + 5, column=12,
                 value=(total_stenter_reprocess_mtd / total_stenter_process_demand_mtd) * 100).alignment = alignment
    else:
        ws6.cell(row=last_filled_row + 5, column=12, value="0").alignment = alignment  # Handle zero division case

    process_demand_data = reprocess_and_process_data['process_demand_data']
    process_demand_start_row = last_filled_row + 2
    print("process_demand_data", process_demand_data)
    if process_demand_data is not None:
        #        process_demand_start_row = last_filled_row + 2
        for row_index, data_dict in enumerate(process_demand_data):
            current_row = process_demand_start_row + row_index
            # Write data into the respective cells
            ws6.cell(row=current_row, column=1, value=data_dict['date_'])
            ws6.cell(row=current_row, column=2, value=data_dict['s_no'])
            ws6.cell(row=current_row, column=3, value=data_dict['article'])
            ws6.cell(row=current_row, column=4, value=data_dict['po_number'])
            ws6.cell(row=current_row, column=5, value=data_dict['k3'])
            ws6.cell(row=current_row, column=6, value=data_dict['meter_production'])

            # Apply border and alignment for each cell in the row
            for col in range(1, 13):
                cell = ws6.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = alignment

    else:
        print("Error: 'process_demand_data' is None or missing in reprocess_and_process_data.")
    #    process_demand_start_row = last_filled_row + 2
    #    for row_index, data_dict in enumerate(process_demand_data):
    #        current_row = process_demand_start_row + row_index
    #        # Write date in column A
    #        ws6.cell(row=current_row, column=1, value=data_dict['date_'])
    #        ws6.cell(row=current_row, column=2, value=data_dict['s_no'])
    #        ws6.cell(row=current_row, column=3, value=data_dict['article'])
    #        ws6.cell(row=current_row, column=4, value=data_dict['po_number'])
    #        ws6.cell(row=current_row, column=5, value=data_dict['k3'])
    #        ws6.cell(row=current_row, column=6, value=data_dict['meter_production'])
    #
    #        for col in range(1, 13):
    #            cell = ws6.cell(row=current_row, column=col)
    #            cell.border = thin_border
    #            cell.alignment = alignment
    print("Indu 2")
    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 1, column=5, value="Total")
    ws6.cell(row=last_filled_row + 1, column=6, value=f"=SUM(F{process_demand_start_row}:F{last_filled_row})")
    calculated_value_process_demand = ws6.cell(row=last_filled_row + 1, column=6).value
    print("calculated_value_process_demand", calculated_value_process_demand)

    for col in range(1, 13):
        cell = ws6.cell(row=last_filled_row + 1, column=col)
        cell.border = thin_border
        cell.alignment = alignment

    last_filled_row = get_last_filled_row(ws6)
    ws6.merge_cells(start_row=last_filled_row + 2, start_column=1, end_row=last_filled_row + 2, end_column=8)
    ws6.cell(row=last_filled_row + 2, column=1, value="RELAX DRYER & HEAT SET  DEMAND")
    for col in range(1, 12):  # Columns A to H (1 to 8)
        cell = ws6.cell(row=last_filled_row + 2, column=col)
        cell.border = thin_border
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='FF5cd65c', end_color='FF5cd65c', fill_type='solid')

        ############################# FILL RELAX DRIER DATA
    last_filled_row = get_last_filled_row(ws6)
    headers = ["Date", "S.No", "Article", "PO Number", "K3", "Meter Production", "Reason", "Deptt."]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws6.cell(row=last_filled_row + 1, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='FF5cd65c', end_color='FF5cd65c', fill_type='solid')

    relax_data = reprocess_and_process_data['relax_dryer_data']
    last_filled_row = get_last_filled_row(ws6)
    relax_data_start_row = last_filled_row + 1
    for row_index, data_dict in enumerate(relax_data):
        if data_dict is None:
            print("Skipping None data entry at index", row_index)
            continue
        current_row = relax_data_start_row + row_index
        ws6.cell(row=current_row, column=1, value=data_dict['date_'])
        ws6.cell(row=current_row, column=2, value=data_dict['s_no'])
        ws6.cell(row=current_row, column=3, value=data_dict['article'])
        ws6.cell(row=current_row, column=4, value=data_dict['po_number'])
        ws6.cell(row=current_row, column=5, value=data_dict['k3'])
        ws6.cell(row=current_row, column=6, value=data_dict['meter_production'])

        for col in range(1, 13):
            cell = ws6.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment

    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 1, column=5, value="Total").alignment = alignment
    ws6.cell(row=last_filled_row + 1, column=6, value=f"=SUM(F{relax_data_start_row}:F{last_filled_row})")
    calculated_value_relax_drier = ws6.cell(row=last_filled_row + 1, column=6).value
    print("calculated_value_relax_drier", calculated_value_relax_drier)

    for col in range(1, 13):
        cell = ws6.cell(row=last_filled_row + 1, column=col)
        cell.border = thin_border
        cell.alignment = alignment

    last_filled_row = get_last_filled_row(ws6)
    ws6.merge_cells(start_row=last_filled_row + 1, start_column=1, end_row=last_filled_row + 1, end_column=5)
    ws6.cell(row=last_filled_row + 1, column=1, value="% Process Demand on Stenter Production")
    ws6.cell(row=last_filled_row + 1, column=1).alignment = alignment
    ws6.cell(row=last_filled_row + 1, column=6, value="0")

    last_filled_row = get_last_filled_row(ws6)
    ws6.merge_cells(start_row=last_filled_row + 1, start_column=1, end_row=last_filled_row + 1, end_column=5)
    ws6.cell(row=last_filled_row + 1, column=1, value="Grand Total").alignment = alignment
    ws6.cell(row=last_filled_row + 1, column=1).alignment = alignment
    total_sum = calculated_value_reprocess_data + calculated_value_process_demand + calculated_value_relax_drier
    print("total_sum:->", total_sum)
    total_sum_formula = f"{calculated_value_reprocess_data}+{calculated_value_process_demand}+{calculated_value_relax_drier}"
    print("total_sum_formula:->", total_sum_formula)
    ws6.cell(row=last_filled_row + 1, column=6, value=total_sum)

    curing_start = last_filled_row + 1
    ws6.cell(row=last_filled_row + 1, column=10, value="CURING EFFECTIVE PRODUCTION").alignment = alignment
    curing_effective_ftd = reprocess_and_process_data['curing_effective_ftd']
    ws6.cell(row=last_filled_row + 1, column=11, value=curing_effective_ftd).alignment = alignment
    curing_effective_mtd = reprocess_and_process_data['curing_effective_mtd']
    ws6.cell(row=last_filled_row + 1, column=12, value=curing_effective_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 2, column=10, value="CURING H/SET").alignment = alignment
    curing_heatsett_ftd = reprocess_and_process_data['curing_heatsett_ftd']
    ws6.cell(row=last_filled_row + 2, column=11, value=curing_heatsett_ftd).alignment = alignment
    curing_heatsett_mtd = reprocess_and_process_data['curing_heatsett_mtd']
    ws6.cell(row=last_filled_row + 2, column=12, value=curing_heatsett_mtd).alignment = alignment

    # last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 3, column=10, value="TOTAL").alignment = alignment
    ws6.cell(row=last_filled_row + 3, column=11, value=curing_effective_ftd + curing_heatsett_ftd).alignment = alignment
    ws6.cell(row=last_filled_row + 3, column=12, value=curing_effective_mtd + curing_heatsett_mtd).alignment = alignment

    stenter_effective_mtd = reprocess_and_process_data.get('stenter_effective_mtd')
    #    if total_sanfo_production_mtd != 0:
    #        process_demand_on_stenter_eff_production = (total_stenter_process_demand_mtd/total_stenter_production_mtd)*100
    #        re_process_on_packing_production = (total_stenter_reprocess_mtd/total_stenter_production_mtd)*100
    #

    if stenter_effective_mtd != 0:
        reprocess_percentage = (total_stenter_reprocess_mtd / stenter_effective_mtd) * 100
        process_demand_percentage = (total_stenter_process_demand_mtd / stenter_effective_mtd) * 100
        process_demand_on_stenter_eff_production = (
                                                           total_stenter_process_demand_mtd / total_stenter_production_mtd) * 100
        re_process_on_packing_production = (total_stenter_reprocess_mtd / total_stenter_production_mtd) * 100

    start_row = get_last_filled_row(ws6) + 1
    ws6.cell(row=start_row, column=8, value="MTD in %").alignment = alignment
    ws6.cell(row=start_row + 2, column=8, value=int(reprocess_percentage)).alignment = alignment
    ws6.cell(row=start_row + 3, column=8, value=int(process_demand_percentage)).alignment = alignment

    titles_with_values = {
        " ": "MTD",
        "STENTER GROSS PRODUCTION": int(total_stenter_production_mtd),
        "RE-PROCESS ( % )": int(total_stenter_reprocess_mtd),
        "PROCESS DEMAND ( % )": int(total_stenter_process_demand_mtd),
        "STENTER EFFECTIVE PRODUCTION": int(stenter_effective_mtd),
        "% Process Demand on Stenter Eff. Production": int(process_demand_on_stenter_eff_production),
        "% Re- Process on Stenter Eff.Production": int(re_process_on_packing_production),
    }

    for title, value in titles_with_values.items():
        last_filled_row = start_row
        merge_end_col = 6 if title.strip() != " " else 7
        ws6.merge_cells(start_row=last_filled_row, start_column=1, end_row=last_filled_row, end_column=merge_end_col)
        ws6.cell(row=last_filled_row, column=1, value=title)
        ws6.cell(row=last_filled_row, column=1).alignment = alignment
        ws6.cell(row=last_filled_row, column=merge_end_col + 1, value=value)  # Fill the next column with the value
        start_row += 1

    ############################# SANFO START HERE
    start_sanfo_reprocess = get_last_filled_row(ws6) + 3
    ws6.merge_cells(start_row=start_sanfo_reprocess, start_column=1, end_row=get_last_filled_row(ws6) + 3,
                    end_column=12)
    ws6.cell(row=start_sanfo_reprocess, column=1,
             value=f"SANFO RE - PROCESS & PROCESS DEMAND REPORT {month_name}-{year}")
    ws6.cell(row=start_sanfo_reprocess, column=1).alignment = alignment
    ws6.cell(row=start_sanfo_reprocess, column=1).fill = PatternFill(start_color='FF5cd65c', end_color='FF5cd65c',
                                                                     fill_type='solid')

    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 3, column=10, value="TOTAL PRODUCTION").alignment = alignment
    total_sanfo_production_ftd = reprocess_and_process_data['total_sanfo_production_ftd']
    ws6.cell(row=last_filled_row + 3, column=11, value=total_sanfo_production_ftd).alignment = alignment
    total_sanfo_production_mtd = reprocess_and_process_data['total_sanfo_production_mtd']
    ws6.cell(row=last_filled_row + 3, column=12, value=total_sanfo_production_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 4, column=10, value="REPROCESS  PRODUCTION").alignment = alignment
    total_sanfo_reprocess_ftd = reprocess_and_process_data['total_sanfo_reprocess_ftd']
    ws6.cell(row=last_filled_row + 4, column=11, value=total_sanfo_reprocess_ftd).alignment = alignment
    total_sanfo_reprocess_mtd = reprocess_and_process_data['total_sanfo_reprocess_mtd']
    ws6.cell(row=last_filled_row + 4, column=12, value=total_sanfo_reprocess_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 5, column=10, value="PERCENTAGE").alignment = alignment
    if total_sanfo_reprocess_ftd != 0:
        ws6.cell(row=last_filled_row + 5, column=11,
                 value=(total_sanfo_production_ftd / total_sanfo_reprocess_ftd) * 100).alignment = alignment
    if total_sanfo_reprocess_mtd != 0:
        ws6.cell(row=last_filled_row + 5, column=12,
                 value=(total_sanfo_production_mtd / total_sanfo_reprocess_mtd) * 100).alignment = alignment

    last_filled_row = get_last_filled_row(ws6)
    sanfo_headers = ["Date", "M/C NO", "Article", "PO Number", "K3", "QTY TOTAL", "Reason", "Deptt."]
    for col_idx, header in enumerate(sanfo_headers, start=1):
        cell = ws6.cell(row=last_filled_row + 1, column=col_idx, value=header)
        cell.border = thin_border
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='FF5cd65c', end_color='FF5cd65c', fill_type='solid')

    ws6.cell(row=get_last_filled_row(ws6) + 1, column=1, value=" REPROCESS DATA")

    sanfo_reprocess_data = reprocess_and_process_data.get('sanfo_reprocess_data')
    sanfo_reprocess_row = get_last_filled_row(ws6) + 1

    for r_idx, data_dict in enumerate(sanfo_reprocess_data):
        current_row = sanfo_reprocess_row + r_idx
        ws6.cell(row=current_row, column=1, value=data_dict['date'])
        ws6.cell(row=current_row, column=2, value=data_dict['mc_number'])
        ws6.cell(row=current_row, column=3, value=data_dict['article'])
        ws6.cell(row=current_row, column=4, value=data_dict['po_number'])
        ws6.cell(row=current_row, column=5, value=data_dict['k3'])
        ws6.cell(row=current_row, column=6, value=data_dict['po_qty'])

    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 2, column=5, value="Total").alignment = alignment
    ws6.cell(row=last_filled_row + 2, column=6, value=f"=SUM(F{sanfo_reprocess_row}:F{last_filled_row})")

    sanfo_process_row = get_last_filled_row(ws6) + 1

    ws6.cell(row=get_last_filled_row(ws6) + 2, column=1, value="PROCESS DEMAND DATA").alignment = alignment

    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 3, column=10, value="TOTAL PRODUCTION").alignment = alignment
    ws6.cell(row=last_filled_row + 3, column=10, value="TOTAL PRODUCTION").alignment = alignment
    total_sanfo_production_ftd = reprocess_and_process_data['total_sanfo_production_ftd']
    ws6.cell(row=last_filled_row + 3, column=11, value=total_sanfo_production_ftd).alignment = alignment
    total_sanfo_production_mtd = reprocess_and_process_data['total_sanfo_production_mtd']
    ws6.cell(row=last_filled_row + 3, column=12, value=total_sanfo_production_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 4, column=10, value="PROCESS DEMAND PRODUCTION").alignment = alignment
    total_sanfo_process_demand_ftd = reprocess_and_process_data['total_sanfo_process_demand_ftd']
    ws6.cell(row=last_filled_row + 4, column=11, value=total_sanfo_process_demand_ftd).alignment = alignment
    total_sanfo_process_demand_mtd = reprocess_and_process_data['total_sanfo_process_demand_mtd']
    ws6.cell(row=last_filled_row + 4, column=12, value=total_sanfo_process_demand_mtd).alignment = alignment

    ws6.cell(row=last_filled_row + 5, column=10, value="PERCENTAGE").alignment = alignment
    if total_sanfo_process_demand_ftd != 0:
        ws6.cell(row=last_filled_row + 5, column=11,
                 value=(total_sanfo_production_ftd / total_sanfo_process_demand_ftd) * 100).alignment = alignment

    if total_sanfo_process_demand_mtd != 0:
        ws6.cell(row=last_filled_row + 5, column=12,
                 value=(total_sanfo_production_mtd / total_sanfo_process_demand_mtd) * 100).alignment = alignment

    sanfo_process_demand_data = reprocess_and_process_data.get('sanfo_process_demand_data')
    sanfo_process_demand_row = get_last_filled_row(ws6) + 1
    for r_idx, data_dict in enumerate(sanfo_process_demand_data):
        current_row = sanfo_process_demand_row + r_idx
        ws6.cell(row=current_row, column=1, value=data_dict['date'])
        ws6.cell(row=current_row, column=2, value=data_dict['mc_number'])
        ws6.cell(row=current_row, column=3, value=data_dict['article'])
        ws6.cell(row=current_row, column=4, value=data_dict['po_number'])
        ws6.cell(row=current_row, column=5, value=data_dict['k3'])
        ws6.cell(row=current_row, column=6, value=data_dict['po_qty'])

    last_filled_row = get_last_filled_row(ws6)
    ws6.cell(row=last_filled_row + 2, column=5, value="Total").alignment = alignment
    ws6.cell(row=last_filled_row + 2, column=6, value=f"=SUM(F{sanfo_process_demand_row}:F{last_filled_row})")
    ws6.cell(row=get_last_filled_row(ws6) + 1, column=5, value="Grand Total").alignment = alignment
    # ws6.cell(row=get_last_filled_row(ws6)+1, column=6, value=).alignment = alignment

    sanfo_effective_mtd = reprocess_and_process_data.get('sanfo_effective_mtd')
    if total_sanfo_production_mtd != 0:
        process_demand_on_stenter_eff_production = (total_sanfo_process_demand_mtd / total_sanfo_production_mtd) * 100
        re_process_on_packing_production = (total_sanfo_reprocess_mtd / total_sanfo_production_mtd) * 100

    if sanfo_effective_mtd != 0:
        reprocess_percentage = (total_sanfo_reprocess_mtd / sanfo_effective_mtd) * 100
        process_demand_percentage = (total_sanfo_process_demand_mtd / sanfo_effective_mtd) * 100

    start_row = get_last_filled_row(ws6) + 1
    ws6.cell(row=start_row, column=8, value="MTD in %").alignment = alignment
    ws6.cell(row=start_row + 2, column=8, value=round(reprocess_percentage, 2)).alignment = alignment
    ws6.cell(row=start_row + 3, column=8, value=round(process_demand_percentage, 2)).alignment = alignment

    titles_with_values = {
        " ": "MTD ",
        "SANFO GROSS PRODUCTION": int(total_sanfo_production_mtd),
        "RE-PROCESS ( % )": int(total_sanfo_reprocess_mtd),
        "PROCESS DEMAND ( % )": int(total_sanfo_process_demand_mtd),
        "SANFO EFFECTIVE PRODUCTION": int(sanfo_effective_mtd),
        "% Process Demand on Packing Production": round(process_demand_on_stenter_eff_production, 2),
        "% Re- Process on Packing Production": round(re_process_on_packing_production, 2)
    }

    for title, value in titles_with_values.items():
        last_filled_row = start_row  # Use the initial start_row value
        merge_end_col = 6 if title.strip() != " " else 7
        ws6.merge_cells(start_row=last_filled_row, start_column=1, end_row=last_filled_row, end_column=merge_end_col)
        ws6.cell(row=last_filled_row, column=1, value=title)
        ws6.cell(row=last_filled_row, column=1).alignment = alignment
        ws6.cell(row=last_filled_row, column=merge_end_col + 1, value=value)
        ws6.cell(row=last_filled_row, column=merge_end_col + 1).alignment = alignment
        start_row += 1

    for row in ws6.iter_rows():
        for cell in row:
            cell.border = thin_border

    ############################         AIRO - 24 REPORT      ################################################
    airo_effective_and_non_effectice_data = await crud.get_airo_effective_noneffective_data(db, date_)
    ws7.cell(row=22, column=1, value=f"AIRO 24 NON-EFFECTIVE PRODUCTION DETAIL AS ON {date_}")
    ws7.cell(row=2, column=8, value=f"DATE {date_}")

    airo_total_ftd = airo_effective_and_non_effectice_data['airo_total_production_ftd']
    ws7.cell(row=3, column=5, value=airo_total_ftd)

    airo_total_mtd = airo_effective_and_non_effectice_data['airo_total_production_mtd']
    ws7.cell(row=3, column=6, value=airo_total_mtd)

    dry_airo_ftd_data = airo_effective_and_non_effectice_data['dry_airo_ftd']
    ws7.cell(row=6, column=5, value=dry_airo_ftd_data)

    dry_airo_mtd_data = airo_effective_and_non_effectice_data['dry_airo_mtd']
    ws7.cell(row=6, column=6, value=dry_airo_mtd_data)

    airo_beat_ftd_data = airo_effective_and_non_effectice_data['airo_beat_ftd']
    ws7.cell(row=7, column=5, value=airo_beat_ftd_data)

    airo_beat_mtd_data = airo_effective_and_non_effectice_data['airo_beat_mtd']
    ws7.cell(row=7, column=6, value=airo_beat_mtd_data)

    dry_airo_production_noneffective_ftd = airo_effective_and_non_effectice_data['dry_airo_production_noneffective_ftd']
    ws7.cell(row=18, column=5, value=dry_airo_production_noneffective_ftd)

    dry_airo_production_noneffective_mtd = airo_effective_and_non_effectice_data['dry_airo_production_noneffective_mtd']
    ws7.cell(row=18, column=6, value=dry_airo_production_noneffective_mtd)

    airo_beat_effective_production_ftd = airo_effective_and_non_effectice_data['airo_beat_effective_production_ftd']
    ws7.cell(row=19, column=5, value=airo_beat_effective_production_ftd)

    airo_beat_effective_production_mtd = airo_effective_and_non_effectice_data['airo_beat_effective_production_mtd']
    ws7.cell(row=19, column=6, value=airo_beat_effective_production_mtd)

    airo_non_effective_data = airo_effective_and_non_effectice_data['airo_non_eff_data']
    airo_non_effect_row = 25
    if airo_non_effective_data is not None:  # Ensure airo_non_effective_data is not None
        for r_idx, data_dict in enumerate(airo_non_effective_data):
            # Check if data_dict is not None
            if data_dict is None:
                continue

            current_row = airo_non_effect_row + r_idx
            ws7.cell(row=current_row, column=1, value=data_dict.get('article', ''))
            ws7.cell(row=current_row, column=2, value=data_dict.get('po_number', ''))
            ws7.cell(row=current_row, column=3, value=data_dict.get('meter_production', ''))
            #            meter_production = data_dict.get('meter_production', '')
            #            ws7.cell(row=current_row, column=3, value=round(meter_production))
            ws7.cell(row=current_row, column=4, value=data_dict.get('finish_Code', ''))
            ws7.cell(row=current_row, column=5, value=data_dict.get('detail', ''))

        for col in ['A', 'B', 'C', 'D', 'E']:
            for row in range(airo_non_effect_row, airo_non_effect_row + len(airo_non_effective_data)):
                cell = ws7[f'{col}{row}']
                cell.border = thin_border
                cell.alignment = alignment

        airo_non_eff_last_fill_row = get_last_filled_row(ws7)

        merge_start_row = airo_non_eff_last_fill_row + 1
        ws7.merge_cells(f'A{merge_start_row}:B{merge_start_row}')
        ws7[f'A{merge_start_row}'].value = "Total"
        ws7[f'A{merge_start_row}'].border = thin_border
        ws7[f'A{merge_start_row}'].alignment = alignment
        ws7[f'A{merge_start_row}'].fill = PatternFill(start_color='FF5cd65c', end_color='FF5cd65c', fill_type='solid')
        ws7[f'B{merge_start_row}'].border = Border(bottom=Side(border_style='thin'))

        ws7[f'C{merge_start_row}'].value = f"=SUM(C25:C{airo_non_eff_last_fill_row})"
        ws7[f'C{merge_start_row}'].border = thin_border
        ws7[f'C{merge_start_row}'].alignment = alignment



    else:
        print("airo_non_effective_data is None")

    airo_running_hours = airo_effective_and_non_effectice_data['airo_running_hours']
    start_row = 6

    for r_idx, data_dict in enumerate(airo_running_hours):
        row_number = start_row + r_idx
        ws7.cell(row=row_number, column=9, value=data_dict['machine'])
        ws7.cell(row=row_number, column=10, value=data_dict.get('A', 0))
        ws7.cell(row=row_number, column=11, value=data_dict.get('B', 0))
        ws7.cell(row=row_number, column=12, value=data_dict.get('C', 0))

    airo_breakdown_hours_data = airo_effective_and_non_effectice_data['airo_breakdown_hours']
    start_row_breakdowns = 9
    for r_idx, data_dict in enumerate(airo_breakdown_hours_data):
        row_number = start_row_breakdowns + r_idx
        ws7.cell(row=row_number, column=9, value=data_dict['mechanical_data'])
        ws7.cell(row=row_number, column=10, value=data_dict['electrical_data'])
        ws7.cell(row=row_number, column=11, value=data_dict['cleaning_data'])
        ws7.cell(row=row_number, column=12, value=data_dict['changeover_data'])
        ws7.cell(row=row_number, column=13, value=data_dict['pm_data'])
        ws7.cell(row=row_number, column=14, value=data_dict['no_program_data'])

    ###########################################################################################################
    ########### FINISHING SECTION PRODUCTION REPORT ###################################################
    print("FINISHING SECTION PRODUCTION REPORT")
    production_data = await crud.get_report_combined_production(db, date_)
    hot_data = await  crud.get_report_hot_and_heat_set_data(db, date_)
    till_date = await  crud.get_report_data_till_date(db, date_)
    heat_till = await crud.get_report_heat_set_data_till_date(db, date_)
    run_duration = await crud.get_report_combined_duration(db, date_)
    # stop_duration = await  crud.get_report_stop_duration_data(db, date_)

    ws1 = wb['Finishing Section Prod']

    ws1['M15'] = heat_till['MTD_Heat_Set']
    # print("heat_till['MTD_Heat_Set']", heat_till['MTD_Heat_Set'])

    df = pd.DataFrame(production_data)
    # print(df)

    hot_df = pd.DataFrame(hot_data)
    # print(hot_df)

    till_df = pd.DataFrame(till_date)

    run_dur_df = pd.DataFrame(run_duration)

    # stop_dur_df = pd.DataFrame(stop_duration)

    machines = ['Stenter-1', 'Stenter-2', 'Stenter-3', 'Stenter-4', 'Stenter-5(HEATSETT)', 'Stenter-5',
                'Stenter-6',
                'Relax-Drier', 'Sanforiser-2', 'Sanforiser-3', 'Sanforiser-4', 'Curing', 'Airo-24'
                ]

    start_row = 15
    start_row1 = 32

    for i, machine in enumerate(machines):
        print(i, machine)

        machine_data = df.get(machine, pd.Series())
        machine_data = machine_data.fillna(0)
        if machine_data is None:
            continue
        print(machine_data)
        row = start_row + i

        for j, (shift, production) in enumerate(machine_data.items()):

            if shift == 'A':
                ws1.cell(row=row, column=2, value=int(production))
            elif shift == 'B':
                ws1.cell(row=row, column=3, value=int(production))
            elif shift == 'C':
                ws1.cell(row=row, column=4, value=int(production))

        hot_data = hot_df.get(machine, pd.Series())
        print(hot_data)
        row = start_row + i
        #        for k, (data, production) in enumerate(hot_data.items()):
        #
        #            if data == 'production_of_Hot_Flue':
        #                ws1.cell(row=row, column=6, value=int(production))
        #            elif data == 'production_of_Heat_Set':
        #                ws1.cell(row=row, column=7, value=int(production))

        for k, (data, production) in enumerate(hot_data.items()):
            if data == 'production_of_Hot_Flue':
                ws1.cell(row=row, column=6, value=int(production) if not pd.isna(production) else 0)
            elif data == 'production_of_Heat_Set':
                ws1.cell(row=row, column=7, value=int(production) if not pd.isna(production) else 0)

        till_data = till_df.get(machine, pd.Series())
        print(till_data)
        row = start_row + i
        for k, (data, production) in enumerate(till_data.items()):
            ws1.cell(row=row, column=8, value=int(production))

        machine_data = run_dur_df.get(machine, pd.Series())
        machine_data = machine_data.fillna(0)
        print("Hello", machine_data)
        row = start_row + i

        for j, (shift, production) in enumerate(machine_data.items()):

            if shift == 'A':
                ws1.cell(row=row, column=17, value=production)
            elif shift == 'B':
                ws1.cell(row=row, column=18, value=production)
            elif shift == 'C':
                ws1.cell(row=row, column=19, value=production)

    # ....................................................................................................
    ws1 = wb['Finishing Section Prod']
    stop_duration = await  crud.get_report_stop_duration_data(db, date_)
    stop_dur_df = pd.DataFrame(stop_duration)

    machines_data = ['Stenter-1', 'Stenter-2', 'Stenter-3', 'Stenter-4', 'Stenter-5',
                     'Stenter-6',
                     'Relax-Drier', 'Sanforiser-2', 'Sanforiser-3', 'Sanforiser-4', 'Curing', 'Airo-24'
                     ]
    start_row1 = 32
    for i, machine in enumerate(machines_data):

        stop_data = stop_dur_df.get(machine, pd.Series())
        print(stop_data)
        row1 = start_row1 + i
        print(start_row1)

        for k, (data, production) in enumerate(stop_data.items()):

            if data == 'mechanical_hours':
                ws1.cell(row=row1, column=2, value=round(production, 2))
            elif data == 'electrical_hours':
                ws1.cell(row=row1, column=3, value=round(production, 2))
            elif data == 'fabric_changeover_hours':
                ws1.cell(row=row1, column=5, value=round(production, 2))
            elif data == 'no_program_hours':
                ws1.cell(row=row1, column=8, value=round(production, 2))
            elif data == 'nip_test_hours':
                ws1.cell(row=row1, column=11, value=round(production, 2))
            elif data == 'startup_for_yardage_hours':
                ws1.cell(row=row1, column=13, value=round(production, 2))
            elif data == 'lead_cloth_burst_hours':
                ws1.cell(row=row1, column=14, value=round(production, 2))
            elif data == 'pm_hours':
                ws1.cell(row=row1, column=6, value=round(production, 2))
            elif data == 'pin_hours':
                ws1.cell(row=row1, column=7, value=round(production, 2))
            elif data == 'air_hours':
                ws1.cell(row=row1, column=10, value=round(production, 2))
            elif data == 'clg_hours':
                ws1.cell(row=row1, column=4, value=round(production, 2))
            elif data == 'misc_hours':
                ws1.cell(row=row1, column=15, value=round(production, 2))
            elif data == 'oil_hours':
                ws1.cell(row=row1, column=12, value=round(production, 2))
            elif data == 'mtd_no_prog_hours':
                ws1.cell(row=row1, column=9, value=round(production, 2))

    ws5 = wb['Monthly MC wise speed']
    data = await  crud.get_production_data_for_year(db, date_)

    for row, (month, values) in enumerate(data.items(), start=3):
        month_abbr = date_.replace(month=datetime.strptime(month, '%B').month, day=1).strftime('%b')
        year = date_.year

        month_year_str = f"{month_abbr} {year}"
        ws5[f'A{row}'] = month_year_str
        ws5[f'B{row}'] = values.get('Stenter-1', {}).get('production_finish', 0)
        ws5[f'C{row}'] = values.get('Stenter-1', {}).get('duration_finish')

        ws5[f'E{row}'] = values.get('Stenter-2', {}).get('production_finish', 0)
        ws5[f'F{row}'] = values.get('Stenter-2', {}).get('duration_finish')

        ws5[f'H{row}'] = values.get('Stenter-5', {}).get('production_finish', 0)
        ws5[f'I{row}'] = values.get('Stenter-5', {}).get('duration_finish')

        ws5[f'K{row}'] = values.get('Stenter-6', {}).get('production_finish', 0)
        ws5[f'L{row}'] = values.get('Stenter-6', {}).get('duration_finish')

    for row, (month, values) in enumerate(data.items(), start=19):
        month_abbr = date_.replace(month=datetime.strptime(month, '%B').month, day=1).strftime('%b')
        year = date_.year

        month_year_str = f"{month_abbr} {year}"
        ws5[f'A{row}'] = month_year_str
        ws5[f'B{row}'] = values.get('Stenter-3', {}).get('production_heat_set', 0)
        ws5[f'C{row}'] = values.get('Stenter-3', {}).get('duration_heat')

        ws5[f'E{row}'] = values.get('Stenter-4', {}).get('production_heat_set', 0)
        ws5[f'F{row}'] = values.get('Stenter-4', {}).get('duration_heat')

        ws5[f'H{row}'] = values.get('Stenter-5', {}).get('production_heat_set', 0)
        ws5[f'I{row}'] = values.get('Stenter-5', {}).get('duration_heat')

        ws5[f'K{row}'] = values.get('Stenter-6', {}).get('production_heat_set', 0)
        ws5[f'L{row}'] = values.get('Stenter-6', {}).get('duration_heat')

    for row, (month, values) in enumerate(data.items(), start=36):
        month_abbr = date_.replace(month=datetime.strptime(month, '%B').month, day=1).strftime('%b')
        year = date_.year

        month_year_str = f"{month_abbr} {year}"
        ws5[f'A{row}'] = month_year_str
        ws5[f'B{row}'] = values.get('Sanforiser-2', {}).get('production', 0)
        ws5[f'C{row}'] = values.get('Sanforiser-2', {}).get('duration_hours')

        ws5[f'E{row}'] = values.get('Sanforiser-3', {}).get('production', 0)
        ws5[f'F{row}'] = values.get('Sanforiser-3', {}).get('duration_hours')

        ws5[f'H{row}'] = values.get('Sanforiser-4', {}).get('production', 0)
        ws5[f'I{row}'] = values.get('Sanforiser-4', {}).get('duration_hours')

    ###################.....................................Heat Set Temp and Width...............................

    #    po_data = await crud.get_po_data_ref111(db, date_)
    #    ws8 = wb["Heat Set Speed & Temp"]
    #
    #    for i, datap in enumerate(po_data):
    #        try:
    #            #print(i, datap)
    #            row = i + 2
    #
    #            ws8[f'A{row}'] = datap['date_']
    #            ws8[f'B{row}'] = datap['machine']
    #            ws8[f'C{row}'] = datap['k1']
    #            ws8[f'D{row}'] = datap['Finish']
    #            ws8[f'E{row}'] = datap['k4']
    #            ws8[f'F{row}'] = datap['po_number']
    #            ws8[f'G{row}'] = datap['po_qty']
    #            ws8[f'H{row}'] = datap['meter']
    #            ws8[f'J{row}'] = datap['finish_glm']
    #            ws8[f'K{row}'] = datap['temperature']
    #            ws8[f'L{row}'] = datap['width']
    #            ws8[f'M{row}'] = datap['speed']
    #
    #            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
    #                for row in range(2, len(po_data) + 2):
    #                    ws8[f'{col}{row}'].border = Border(
    #                        left=Side(border_style='thin'),
    #                        right=Side(border_style='thin'),
    #                        top=Side(border_style='thin'),
    #                        bottom=Side(border_style='thin')
    #                    )
    #                    ws8[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    #
    #        except Exception as e:
    #           print(e)

    path_ = os.path.join(dirname, f"Reports/Vardhman_Report_{date_}.xlsx")
    wb.save(path_)

    return path_


##.......................Dyeing Section Report....................


##For testing purpose only
STENTER_LIST = ["Stenter-1", "Stenter-2", "Stenter-3", "Stenter-4"]

PAD_DRY_LIST = ["Pad Dry-1", "Pad Dry-2", "Pad Dry-3"]
CPB_LIST = ["CPB-1", "CPB-2"]
PAD_STEAM_LIST = ["KPS", "OPS"]
THERMOFIX_LIST = ["Thermofix"]
JIGGER_LIST = ["Jigger-1", "Jigger-2", "Jigger-3"]
JET_LIST = ["Sclavos-1", "Sclavos-2", "Sclavos-3", "Sclavos-4"]


async def update_monthly_prod_sheet(wb, sheet_name, po_data, date_, year, center_alignment):
    ws = wb[sheet_name]
    month_name = date_.strftime("%B")
    ws["A2"].value = f"MONTH: {month_name}.{year}"
    for i, datap in enumerate(po_data):
        try:
            row = i + 4
            ws[f'A{row}'] = datap['date']
            ws[f'B{row}'] = datap['machine']
            ws[f'C{row}'] = datap['operation_name']
            ws[f'D{row}'] = datap['k1']
            ws[f'E{row}'] = datap['k3']
            ws[f'F{row}'] = datap['k4']
            ws[f'G{row}'] = datap['po_number']
            ws[f'H{row}'] = datap['total_meters']
            # ws[f'I{row}'] = datap.get('KGS')
            ws[f'J{row}'] = datap['greige_glm']
            ws[f'K{row}'] = datap['total_hours']
        except Exception as e:
            print(f"Error processing row {i + 4}: {e}")
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        for row in range(4, len(po_data) + 4):
            ws[f'{col}{row}'].alignment = center_alignment


async def generate_dyeing_report_data(db: Session, date_: date):
    wb = load_workbook(os.path.join(dirname, "Reports/Template/AT_DYEING_SECTION_REPORT.xlsx"))
    ws = wb["Machine Wise Production"]
    machine_data = [
        (PAD_DRY_LIST, "A", "B"),
        (CPB_LIST, "C"),
        (PAD_STEAM_LIST, "D"),
        (THERMOFIX_LIST, "E"),
        (JIGGER_LIST, "F"),
        (JET_LIST, "G")
    ]
    start_row = 6
    for machine_list, *columns in machine_data:
        production_data = await dyeing_report.get_machine_production_for_month(db, date_, machine_list)
        machine_names = list(production_data.keys())
        production_values = list(production_data.values())
        for i in range(len(machine_names)):
            if len(columns) == 2:
                ws[f"{columns[0]}{start_row + i}"] = machine_names[i]
                ws[f"{columns[1]}{start_row + i}"] = production_values[i]
            else:
                ws[f"{columns[0]}{start_row + i}"] = production_values[
                    i]

    ###........................Pad Steamer Prod.....................................................

    ws1 = wb["Pad Steamer Prod"]

    def populate_cells(ws, data, row_start, col_map):
        row = row_start
        for date, values in data.items():
            ws.cell(row=row, column=1).value = date  # Date in first column
            for key, col in col_map.items():
                ws.cell(row=row, column=col).value = values.get(key, 0)
            row += 1

    col_mappings = {
        'ops_prod': {'VATDEV': 2, 'CPB': 3, 'ECOWASH': 4},
        'kps_prod': {'VATDEV': 5, 'CPB': 6, 'ECOWASH': 7},
        'relax_prod': {'CPB': 8, 'ECOWASH': 9},
        'kusters_prod': {'CPB': 10, 'ECOWASH': 11},
        'ops_strip': {'STRIPPNG': 21},
        'kps_strip': {'STRIPPNG': 22},
        'ops_rc': {'RDUCCLR': 24},
        'kps_rc': {'RDUCCLR': 25},
        'ops_2wash': {'SOAPING': 27},
        'kps_2wash': {'SOAPING': 28},
        'relax_2wash': {'SOAPING': 29},
        'kuster_2wash': {'SOAPING': 30},
        'ops_double': {'SOAPING': 13},
        'kps_double': {'SOAPING': 14},
        'ops_redye': {'REPADDIN': 16},
        'kps_redye': {'REPADDIN': 17},
        'relax_redye': {'REPADDIN': 18},
        'kusters_redye': {'REPADDIN': 19},
    }

    ops_prod = await dyeing_report.get_machine_op_production_for_month(db, date_, 'OPS')
    kps_prod = await dyeing_report.get_machine_op_production_for_month(db, date_, 'KPS')
    relax_prod = await dyeing_report.get_machine_op_production_for_month(db, date_, 'Relax-Washer')
    kusters_prod = await dyeing_report.get_machine_op_production_for_month(db, date_, 'Kuster Washer')

    ops_strip = await dyeing_report.get_production_till_date(db, date_, 'OPS', 'STRIPPNG')
    kps_strip = await dyeing_report.get_production_till_date(db, date_, 'KPS', 'STRIPPNG')
    ops_rc = await dyeing_report.get_production_till_date(db, date_, 'OPS', ['RDUCCLR'])

    kps_rc = await dyeing_report.get_production_till_date(db, date_, 'KPS', ['RDUCCLR'])
    ops_2wash = await dyeing_report.get_production_till_date(db, date_, 'OPS', ['SOAPING'])
    kps_2wash = await dyeing_report.get_production_till_date(db, date_, 'KPS', ['SOAPING'])
    relax_2wash = await dyeing_report.get_production_till_date(db, date_, 'Relax-Washer',
                                                               ['SOAPING'])
    kuster_2wash = await dyeing_report.get_production_till_date(db, date_, 'Kusters-Washer',
                                                                ['SOAPING'])
    ops_dd = await dyeing_report.get_production_till_date(db, date_, 'OPS', 'SOAPING')
    kps_dd = await dyeing_report.get_production_till_date(db, date_, 'KPS', 'SOAPING')
    ops_rd = await dyeing_report.get_production_till_date(db, date_, 'OPS', ['REPADDIN'])
    kps_rd = await dyeing_report.get_production_till_date(db, date_, 'KPS', ['REPADDIN'])
    relax_rd = await dyeing_report.get_production_till_date(db, date_, 'Relax-Washer', ['REPADDIN'])
    kusters_rd = await dyeing_report.get_production_till_date(db, date_, 'Kusters-Washer', ['REPADDIN'])

    populate_cells(ws1, ops_prod, 5, col_mappings['ops_prod'])
    populate_cells(ws1, kps_prod, 5, col_mappings['kps_prod'])
    populate_cells(ws1, relax_prod, 5, col_mappings['relax_prod'])
    populate_cells(ws1, kusters_prod, 5, col_mappings['kusters_prod'])
    populate_cells(ws1, ops_strip, 5, col_mappings['ops_strip'])
    populate_cells(ws1, kps_strip, 5, col_mappings['kps_strip'])
    populate_cells(ws1, ops_rc, 5, col_mappings['ops_rc'])
    populate_cells(ws1, kps_rc, 5, col_mappings['kps_rc'])
    populate_cells(ws1, ops_2wash, 5, col_mappings['ops_2wash'])
    populate_cells(ws1, kps_2wash, 5, col_mappings['kps_2wash'])
    populate_cells(ws1, relax_2wash, 5, col_mappings['relax_2wash'])
    populate_cells(ws1, kuster_2wash, 5, col_mappings['kuster_2wash'])
    populate_cells(ws1, ops_dd, 5, col_mappings['ops_double'])
    populate_cells(ws1, kps_dd, 5, col_mappings['kps_double'])
    populate_cells(ws1, ops_rd, 5, col_mappings['ops_redye'])
    populate_cells(ws1, kps_rd, 5, col_mappings['kps_redye'])
    populate_cells(ws1, relax_rd, 5, col_mappings['relax_redye'])
    populate_cells(ws1, kusters_rd, 5, col_mappings['kusters_redye'])

    ##...................................JET AND JIGGER PROD..................................
    ws2 = wb["JET & JIGGER Prod"]

    async def populate_machine_data(sheet, start_row, machine_data, fetch_data_func, db, date_, date_column="A"):
        for machine_list, *columns in machine_data:
            production_data = await fetch_data_func(db, date_, machine_list)
            date_name = list(production_data.keys())
            production_values = list(production_data.values())
            for i in range(len(date_name)):
                sheet[f"{date_column}{start_row + i}"] = date_name[i]
                if len(columns) == 2:
                    sheet[f"{columns[0]}{start_row + i}"] = production_values[i]
                else:
                    sheet[f"{columns[0]}{start_row + i}"] = production_values[i]

    machine_data_production = [
        (JIGGER_LIST, "B"),
        (JET_LIST, "E")
    ]
    machine_data_hour = [
        (JIGGER_LIST, "C"),
        (JET_LIST, "F")
    ]
    await populate_machine_data(ws2, 4, machine_data_production, dyeing_report.get_machine_production_till_date, db,
                                date_, date_column="A")
    await populate_machine_data(ws2, 4, machine_data_hour, dyeing_report.get_machine_duration_till_date, db, date_,
                                date_column="D")

    year = date_.year
    center_alignment = Alignment(horizontal='center', vertical='center')
    ##.............................................JET MONTHLY PROD...............................
    jet_data = await dyeing_report.get_jet_production(db, date_, JET_LIST)
    await update_monthly_prod_sheet(wb, "JET Monthly Prod", jet_data, date_, year, center_alignment)

    ##........................................JIGGER MONTHLY PROD..................................
    jigger_data = await dyeing_report.get_jet_production(db, date_, JIGGER_LIST)
    await update_monthly_prod_sheet(wb, "JIGGER Monthly Prod", jigger_data, date_, year, center_alignment)

    ##........................................PAD DYEING PROD SHEET..................................
    ws5 = wb["PAD Dyeing Prod"]
    await fill_pad_dyeing_sheet(ws5, db, date_)

    ##....................................PAD DRY PRODUCTION SHEET......................................
    ws6 = wb['Pad Dry Prod']
    await fill_pad_dyeing_report_sheet(ws6, db, date_)

    ##........................................DYEING REPORT SHEET..................................
    ws7 = wb["Dyeing Report"]
    ws7.cell(row=2, column=2, value=f" WIP DYEING AS ON {date_}")
    ws7.cell(row=2, column=9, value=f" DYEING MACHINE PRODUCTION {date_}")
    await fill_dyeing_report_sheet(ws7, db, date_)

    path_ = os.path.join(dirname, f"Reports/Vardhman_Dyeing_Report_{date_}.xlsx")
    wb.save(path_)

    return path_


async def fill_pad_dyeing_sheet(ws5, db, date_):
    # Fetch monthly production data
    monthly_production_pad_dying = await dyeing_report.get_monthly_production_of_pad_dying_sheet(db, date_)

    # Machines and column mappings
    machines = ["Pad Dry-1", "Pad Dry-2", "Pad Dry-3", "CPB-1", "CPB-2"]
    columns = {
        "effective_production_value": 0,
        "disperse_non_effective_production_value": 6,
        "double_dyeing_non_effective_production_value": 12,
        "re_dyeing_non_effective_production_value": 18
    }
    base_columns = ["B", "C", "D", "E", "F"]

    # Generate column mappings dynamically
    column_mappings = {}
    for i, machine in enumerate(machines):
        column_mappings[machine] = {}
        for key, offset in columns.items():
            column_mappings[machine][chr(ord(base_columns[i]) + offset)] = key

    # Dictionary to map dates to rows
    date_row_mapping = {}

    # Iterate through the production data and fill the worksheet
    for machine, data_list in monthly_production_pad_dying.items():
        columns = column_mappings.get(machine, {})
        for entry in data_list:
            date = entry["date"]

            # If the date hasn't been mapped to a row yet, assign a new row
            if date not in date_row_mapping:
                row_num = len(date_row_mapping) + 6  # Starting row number (adjust as necessary)
                date_row_mapping[date] = row_num
                ws5[f"A{row_num}"] = date  # Set the date in column A

            row_num = date_row_mapping[date]

            # Fill the row with values based on the column mapping
            for col, value_key in columns.items():
                ws5[f"{col}{row_num}"] = entry.get(value_key, 0)  # Default to 0 if value is missing

    return None


async def fill_pad_dyeing_report_sheet(ws6, db, date_):
    pad_dry_production_sheet_data = await dyeing_report.get_pad_dry_whole_production(db, date_)
    for row, entry in enumerate(pad_dry_production_sheet_data, start=2):
        date_str = entry['date_']
        stop_time_ = entry['stop_time']
        work_center_code = entry['work_center_code']
        machine_code = entry['machine_code']
        machine = entry['machine']
        operation_name = entry['operation_name']
        article = entry['article']
        variants_data = entry['variants_data']
        finish = entry['finish']
        shade_no = entry['shade_no']
        po_number = entry['po_number']
        shift_data = entry['shift_data']
        meter_production = entry['meter_production']
        speed = entry['speed']
        run_duration = entry['run_duration']
        glm = entry['glm']
        work_center_name = entry['work_center_name']

        ws6[f'A{row}'] = date_str
        ws6[f'B{row}'] = stop_time_
        ws6[f'G{row}'] = work_center_code
        ws6[f'H{row}'] = work_center_name
        ws6[f'I{row}'] = machine_code
        ws6[f'J{row}'] = machine
        ws6[f'K{row}'] = operation_name
        ws6[f'S{row}'] = article
        ws6[f'T{row}'] = variants_data
        ws6[f'U{row}'] = finish
        ws6[f'V{row}'] = shade_no
        ws6[f'Y{row}'] = po_number
        ws6[f'AA{row}'] = shift_data
        ws6[f'AH{row}'] = meter_production
        ws6[f'AI{row}'] = speed
        ws6[f'AJ{row}'] = run_duration
        ws6[f'AR{row}'] = glm

    return None


async def fill_dyeing_report_sheet(ws7, db, date_):
    report_data = await dyeing_report.get_monthly_production_of_dying_report_ftd_and_mtd_data(db=db, date_=date_)
    fresh_production_today = report_data['fresh_production_today']
    ws7.cell(row=4, column=11, value=fresh_production_today).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
    fresh_production_mtd = report_data['fresh_production_mtd']
    ws7.cell(row=4, column=12, value=fresh_production_mtd).alignment = Alignment(horizontal='center', vertical='center')

    re_dyeing_today = report_data['re_dyeing_today']
    ws7.cell(row=5, column=11, value=re_dyeing_today).alignment = Alignment(horizontal='center', vertical='center')
    re_dyeing_mtd = report_data['re_dyeing_mtd']
    ws7.cell(row=5, column=12, value=re_dyeing_mtd).alignment = Alignment(horizontal='center', vertical='center')

    disperse_padding_today = report_data['disperse_padding_today']
    ws7.cell(row=6, column=11, value=disperse_padding_today).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
    disperse_padding_mtd = report_data['disperse_padding_mtd']
    ws7.cell(row=6, column=12, value=disperse_padding_mtd).alignment = Alignment(horizontal='center', vertical='center')

    ops_production_today = report_data['ops_production_today']
    ws7.cell(row=7, column=11, value=ops_production_today).alignment = Alignment(horizontal='center', vertical='center')
    ops_production_mtd = report_data['ops_production_mtd']
    ws7.cell(row=7, column=12, value=ops_production_mtd).alignment = Alignment(horizontal='center', vertical='center')

    kps_production_today = report_data['kps_production_today']
    ws7.cell(row=8, column=11, value=kps_production_today).alignment = Alignment(horizontal='center', vertical='center')
    kps_production_mtd = report_data['kps_production_mtd']
    ws7.cell(row=8, column=12, value=kps_production_mtd).alignment = Alignment(horizontal='center', vertical='center')

    relax_washer_production_today = report_data['relax_washer_production_today']
    ws7.cell(row=9, column=11, value=relax_washer_production_today).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
    relax_washer_production_mtd = report_data['relax_washer_production_mtd']
    ws7.cell(row=9, column=12, value=relax_washer_production_mtd).alignment = Alignment(horizontal='center',
                                                                                        vertical='center')

    thermofix_production_today = report_data['thermofix_production_today']
    ws7.cell(row=10, column=11, value=thermofix_production_today).alignment = Alignment(horizontal='center',
                                                                                        vertical='center')
    thermofix_production_mtd = report_data['thermofix_production_mtd']
    ws7.cell(row=10, column=12, value=thermofix_production_mtd).alignment = Alignment(horizontal='center',
                                                                                      vertical='center')

    jiger_production_today = report_data['jiger_production_today']
    ws7.cell(row=11, column=11, value=jiger_production_today).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
    jiger_production_mtd = report_data['jiger_production_mtd']
    ws7.cell(row=11, column=12, value=jiger_production_mtd).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')

    jet_production_today = report_data['jet_production_today']
    ws7.cell(row=12, column=11, value=jet_production_today).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
    jet_production_mtd = report_data['jet_production_mtd']
    ws7.cell(row=12, column=12, value=jet_production_mtd).alignment = Alignment(horizontal='center', vertical='center')

    double_dyeing_production_today = report_data['double_dyeing_production_today']
    ws7.cell(row=15, column=10, value=double_dyeing_production_today).alignment = Alignment(horizontal='center',
                                                                                            vertical='center')
    double_dyeing_production_mtd = report_data['double_dyeing_production_mtd']
    ws7.cell(row=16, column=10, value=double_dyeing_production_mtd).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')

    monthly_production = report_data['monthly_production']
    for r_idx, (date_str, entry) in enumerate(monthly_production.items(), start=22):
        date_str = date_str
        date_ = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
        vat_production = entry['PCVATPAD']
        pd_chemical_production = entry['CHEMPAD']
        double_dyeing_production = entry['DBDYEING']
        cpb_production = entry['CPBPAD']
        e_control = entry['ECONTROL']
        pc_single_bath = entry['DISVTPAD']
        # disperse_padding = entry['0']
        re_dyeing = entry['REPADDIN']

        ws7[f'B{r_idx}'] = date_
        ws7[f'C{r_idx}'] = vat_production
        ws7[f'D{r_idx}'] = pd_chemical_production
        ws7[f'E{r_idx}'] = double_dyeing_production
        ws7[f'F{r_idx}'] = cpb_production
        ws7[f'G{r_idx}'] = e_control
        ws7[f'H{r_idx}'] = pc_single_bath
        # ws7[f'I{r_idx}'] = disperse_padding
        ws7[f'J{r_idx}'] = re_dyeing

    stopage_production = report_data['stopage_data']
    machine_row_mapping = {
        "Pad Dry-1": 16,
        "Pad Dry-2": 17,
        "Pad Dry-3": 18
    }
    for entry in stopage_production:
        machine_name = entry['machine']
        row = machine_row_mapping.get(machine_name)
        if row:
            no_prg_production = entry['no_prg']
            electrical_data_production = entry['electrical_data']
            mechanical_data_production = entry['mechanical_data']
            mc_maintenance_production = entry['mc_maintenance']
            major_concern_production = entry['major_concern']
            any_process_abnomality_deviation = entry['any_process_abnomality_deviation']

            ws7[f'L{row}'] = no_prg_production
            ws7[f'M{row}'] = electrical_data_production
            ws7[f'N{row}'] = mechanical_data_production
            ws7[f'P{row}'] = mc_maintenance_production
            ws7[f'R{row}'] = major_concern_production
            ws7[f'T{row}'] = any_process_abnomality_deviation

    utility_data = report_data['utility_data']
    utility_machine_cell_mapping = {
        "Pad Dry": "O4",
        "CPB": "O5",
        "OPS": "N7",
        "KPS": "N8",
        "Relax-Washer": "N9",
        "Kusters Washer": "N10",
        "Thermofix": "N11",
        "Jigger": "N12",
        "Jet": "N13"
    }
    for entry in utility_data:
        machine_name = entry['machine_list']
        total_utility = entry['total_utility']
        cell = utility_machine_cell_mapping.get(machine_name)

        if cell:
            ws7[cell] = total_utility

    return None


##............................Preparatory Report.....................................


async def write_kg_production_data(new_sheet, start_row, machine, kg_production_data, data_column: str):
    # Ensure kg_production_data is not empty
    if kg_production_data:
        for date_str, production_value in kg_production_data.items():
            row_index = start_row + list(kg_production_data.keys()).index(date_str)
            new_sheet[f"{data_column}{row_index}"] = production_value
    else:
        print(f"No production data found for machine: {machine} on date.")


async def populate_machine_data_with_production(db, new_sheet, start_row, machine, date_, data_column="G"):
    # Fetch the production data in KG for the machine
    kg_production_data = await preparatory_report.calculate_production_in_kg(db, date_, machine)

    # Fill the production data in the specified column
    await write_kg_production_data(new_sheet, start_row, machine, kg_production_data, data_column)


async def populate_machine_data(sheet, start_row, machine_data, fetch_data_func, db, date_, date_column="A"):
    for machine_list, *columns in machine_data:
        production_data = await fetch_data_func(db, date_, machine_list)
        date_name = list(production_data.keys())
        production_values = list(production_data.values())
        for i in range(len(date_name)):
            sheet[f"{date_column}{start_row + i}"] = date_name[i]
            if len(columns) == 2:
                sheet[f"{columns[0]}{start_row + i}"] = production_values[i]
            else:
                sheet[f"{columns[0]}{start_row + i}"] = production_values[i]


def populate_cells(ws, data, row_start, col_map):
    row = row_start
    for date, values in data.items():
        ws.cell(row=row, column=1).value = date  # Date in first column
        for key, col in col_map.items():
            ws.cell(row=row, column=col).value = values.get(key, 0)
        row += 1


async def generate_preparatory_report_data_old(db: Session, date_: date):
    wb = load_workbook(os.path.join(dirname, "Reports/Template/Preparatory_Report_Template.xlsx"))
    # List of machines
    machines_temp2 = ['Osthoff-1', 'Osthoff-2', 'PTR', 'NPS', 'Batcher-1', 'Batcher-2', 'Perble',
                      'Lafer-1', 'Lafer-2', 'Lafer-3', 'Lafer-4', 'Lafer-5-Raising',
                      'Lafer-5-Shearing', 'Soaper', 'Swastik', 'Merceriser-1', 'Merceriser-2', 'Xetma-1', 'Xetma-2']
    #
    # Dictionary of column mappings for the different calculations
    col_mappings = {
        'process_demand': {'Process Demand': 3},
        'rep_prod': {'Reprocess': 4}

    }
    col_dur = {
        'mech_dur': {'Mechanical Breakdown': 11},
        'elec_dur': {'Electrical Breakdown': 12},
        'mach_clean': {'Machine cleaning': 13},
        'changeover': {'Fabric changeover': 14},
        'pm': {'Preventive Maintenance': 15},
        'no_power': {'No Power': 16},
        'no_steam': {'No Steam': 17},
        'no_program': {'No Program': 18},
        'man_power_shortage': {'Man power shortage': 19},
        'insect_problem': {'Insect Problem': 21},
        'no_trolley': {'No Trolley': 22}
    }
    current_start_row = 4

    # For each machine, copy the template, rename it, and apply the data population logic
    for machine in machines_temp2:
        print(machine)
        # Copy "temp" sheet and rename it with the machine name
        new_sheet = wb.copy_worksheet(wb["temp2"])
        new_sheet.title = machine
        new_sheet['A1'].value = f'{machine} Production'
        new_sheet['J1'].value = f"{date_.strftime('%B')} {date_.year}"
        # data = preparatory_report.get_per_kg_data_for_month
        # print(data)

        # Machine-specific data
        machine_data_production = [(machine, "B")]
        machine_data_duration = [(machine, "I")]
        production_till_date = [(machine, "F")]
        dur_till_date = [(machine, "J")]
        production_in_kg = [(machine, "G")]
        cumulative_production = [(machine, "H")]

        await populate_machine_data(new_sheet, 4, machine_data_production,
                                    preparatory_report.fetch_machine_production_for_month, db, date_)

        await populate_machine_data(new_sheet, 4, production_in_kg,
                                    preparatory_report.fetch_prod_kgs_machine_production_for_month, db, date_)

        await populate_machine_data(new_sheet, 4, machine_data_duration,
                                    preparatory_report.fetch_machine_duration_for_month, db, date_)

        await populate_machine_data(new_sheet, 4, production_till_date,
                                    preparatory_report.fetch_machine_production_till_date, db, date_)

        await populate_machine_data(new_sheet, 4, dur_till_date, preparatory_report.fetch_machine_duration_till_date,
                                    db, date_)

        # populate_machine_data(new_sheet, 4, cumulative_production,
        #                           preparatory_report.fetch_mtd_prod_kgs_machine_production_till_date, db, date_)

        # await populate_machine_data(new_sheet, 4, production_in_kg, preparatory_report.get_per_kg_data_for_month,
        #                          db, date_)

        # await populate_machine_data(new_sheet, 4, cumulative_production,
        #                            preparatory_report.calculate_production_in_kg_till_date,
        #                           db, date_)

        ##.......................

        stop_list = ['Machanical Breakdown', 'Electrical Breakdown', 'Machine cleaning', 'Fabric changeover',
                     'Preventive Maintenance', 'No Power', 'No Steam', 'Man Power Shortage', 'No Program',
                     'Insect Problem', 'No Trolley']

        month_production = await preparatory_report.fetch_misc_duration_daily(db, date_, machine, stop_list)

        start_row = 4
        column = 20

        for i, (date_str, value) in enumerate(month_production.items(), start=start_row):
            new_sheet.cell(row=i, column=column).value = value

        rep_prod = await preparatory_report.fetch_operation_production_for_month(db, date_, machine, 'Reprocess')
        populate_cells(new_sheet, rep_prod, 4, col_mappings['rep_prod'])

        process_demand = await preparatory_report.fetch_operation_production_for_month(db, date_, machine,
                                                                                       'Process Demand')
        populate_cells(new_sheet, process_demand, 4, col_mappings['process_demand'])

        mech_dur = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                               'Machanical Breakdown')
        populate_cells(new_sheet, mech_dur, 4, col_dur['mech_dur'])

        elec_dur = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                               'Electrical Breakdown')
        populate_cells(new_sheet, elec_dur, 4, col_dur['elec_dur'])

        mach_clean = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'Machine cleaning')
        populate_cells(new_sheet, mach_clean, 4, col_dur['mach_clean'])

        changeover = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                                 'Fabric changeover')
        populate_cells(new_sheet, changeover, 4, col_dur['changeover'])

        pm = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'Preventive Maintenance')
        populate_cells(new_sheet, pm, 4, col_dur['pm'])

        no_power = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Power')
        populate_cells(new_sheet, no_power, 4, col_dur['no_power'])

        no_steam = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Steam')
        populate_cells(new_sheet, no_steam, 4, col_dur['no_steam'])

        no_program = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Program')
        populate_cells(new_sheet, no_program, 4, col_dur['no_program'])

        man_power_shortage = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                                         'Man Power Shortage')
        populate_cells(new_sheet, man_power_shortage, 4, col_dur['man_power_shortage'])

        insect_problem = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                                     'Insect Problem')
        populate_cells(new_sheet, insect_problem, 4, col_dur['insect_problem'])

        no_trolley = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Trolley')
        populate_cells(new_sheet, no_trolley, 4, col_dur['no_trolley'])

    if "temp2" in wb.sheetnames:
        temp_sheet = wb["temp2"]
        wb.remove(temp_sheet)

    if "temp" in wb.sheetnames:
        temp_sheet = wb["temp"]
        wb.remove(temp_sheet)

    path_ = os.path.join(dirname, f"Reports/Vardhman_Preparartory_Report_{date_}.xlsx")

    wb.save(path_)
    print("save wb")

    return path_


async def download_preparatory_report_data(db: Session, date_: date):
    path_ = await generate_preparatory_report_data(db, date_)
    print("path", path_)
    # path_ = os.path.join(dirname, f"Reports/Vardhman_Preparartory_Report_{date_}.xlsx")

    new_wb = load_workbook(path_)

    sheet_names = new_wb.sheetnames
    print("Loaded sheets:", sheet_names)

    exclude_sheets = ['FRONT PAGE', 'COST SHEET   (2)']

    for sheet_name in sheet_names:
        if sheet_name in exclude_sheets:
            print("Skipping sheet:", sheet_name)
            continue
        sheet = new_wb[sheet_name]
        machine = sheet_name
        print("machine is :->", machine)
        try:
            kg_production_data = await preparatory_report.calculate_production_in_kg(db, date_, machine)
            kg_production_data_till_date = await preparatory_report.calculate_production_in_kg_till_date(db, date_,
                                                                                                         machine)
        except Exception as e:
            print(f"Error occurred: {e}")
            db.rollback()  # Rollback the transaction
            return None  # Or handle the error appropriately

        #        kg_production_data = await preparatory_report.calculate_production_in_kg(db, date_, machine)
        #        print("kg_production_data:->",kg_production_data)
        #        kg_production_data_till_date = await preparatory_report.calculate_production_in_kg_till_date(db, date_, machine)
        #        print("kg_production_data_till_date:->",kg_production_data_till_date)

        start_row = 4
        for index, date_str in enumerate(kg_production_data):
            row_number = start_row + index
            sheet[f'G{row_number}'] = kg_production_data.get(date_str)

        for index, date_str in enumerate(kg_production_data_till_date):
            row_number = start_row + index
            sheet[f'H{row_number}'] = kg_production_data_till_date.get(date_str)

    path_ = os.path.join(dirname, f"Reports/Vardhman_Preparartory_Report_{date_}.xlsx")
    new_wb.save(path_)
    return path_


#####################  PREPARATORY PRODUCTION REPORT HERE
def fill_machine_data(row_number, ws, machine_data):
    for record in machine_data:
        # Fill in the machine name
        ws.cell(row=row_number, column=1, value=record['machine'])

        # Fill in the production for shifts A, B, and C
        ws.cell(row=row_number, column=2, value=record.get('A', 0))
        ws.cell(row=row_number, column=3, value=record.get('B', 0))
        ws.cell(row=row_number, column=4, value=record.get('C', 0))
        ws.cell(row=row_number, column=9, value=record.get('reprocess_data', 0))
        ws.cell(row=row_number, column=10, value=record.get('process_demand_data', 0))
        ws.cell(row=row_number, column=15, value=record.get('running_hours', 0))
        row_number += 1


async def write_machine_data_to_worksheet(ws, pr_production_data, pr_machine_columns, start_row):
    for machine_info in pr_production_data:
        machine_name = machine_info["machine"]
        production_data = machine_info["production_data"]
        glm_data = machine_info["glm_data"]

        if machine_name in pr_machine_columns:
            production_col, glm_col = pr_machine_columns[machine_name]
            for i, shift in enumerate(['A', 'B', 'C']):
                ws.cell(row=start_row + i, column=production_col, value=production_data[shift])  # Production
                ws.cell(row=start_row + i, column=glm_col, value=glm_data[shift])  # GLM Data


async def fill_stopage_data(ws4, stopage_data, column_mapping_stopage_data):
    start_row_of_stopage_data = 48
    for machine, downtimes in stopage_data.items():
        # if machine == "machine_durations":
        #     continue
        # Set machine name in column A (first column)
        ws4[f"A{start_row_of_stopage_data}"] = machine

        # Iterate over each downtime type and fill the corresponding column
        for downtime, value in downtimes.items():
            if downtime in column_mapping_stopage_data:
                col = column_mapping_stopage_data[downtime]
                ws4[f"{col}{start_row_of_stopage_data}"] = value

        # Move to the next row for the next machine
        start_row_of_stopage_data += 1


async def get_preparatory_production_report_data_old(db: Session, date_: date):
    wb = load_workbook(os.path.join(dirname, f"Reports/Template/Preparatory_Production_Report_Template.xlsx"))
    ws1 = wb["MTD PROD KGS & RE MTRS_"]
    ws2 = wb["MTD PRODUTIN"]
    ws3 = wb["EFFECTIN PROB."]
    ws4 = wb["FRONT PAGE "]
    #    ws5 = wb["PR"]
    ws6 = wb["NPS PROD"]
    ws7 = wb["MR OLD"]
    ws8 = wb["MR. NEW "]

    ##### ******************************* KG PRODUCTION REPORT SHEET
    production_data = await get_production_kg_of_each_machine(db, date_)
    start_row = 5
    date_range = pd.date_range(datetime(date_.year, date_.month, 1),
                               datetime(date_.year, date_.month, calendar.monthrange(date_.year, date_.month)[1]))

    for row_index, single_date in enumerate(date_range, start=start_row):
        ws1[f"A{row_index}"] = single_date.strftime('%d-%b-%y')
        ws1[f"W{row_index}"] = single_date.strftime('%d-%b-%y')

    kg_production_column_mappings = {'SANDO': 2, 'OSTHOFF-1': 3, 'OSTHOFF-2': 4, 'PTR': 5, 'NPR': 6, 'OMR': 7, 'NMR': 8,
                                     'NPS': 9, 'Batcher-1': 10, 'Batcher-2': 11, 'OSTHOFF-1 BATCH': 12, 'Xetma-1': 13,
                                     'Xetma-2': 14,
                                     'LAFFER': 15, 'LAFFER-2': 16, 'LAFFER-3': 17, 'LAFFER-4': 18, 'LAFFER RASING': 19,
                                     'LAFFER SHARING': 20, 'SOAPER-2': 21
                                     }
    for machine_data in production_data:
        for machine, production_dict in machine_data.items():
            column = kg_production_column_mappings.get(machine)
            if column:
                for row_index, single_date in enumerate(date_range, start=start_row):
                    production_kg = production_dict.get(single_date.strftime('%Y-%m-%d'), 0)
                    ws1.cell(row=row_index, column=column, value=production_kg)

    # ....................................... REPROCESS MTR PRODUCTION
    reprocess_production_column_mappings = {'PTR': 24, 'NPR': 25, 'Merceriser-1': 26, 'Merceriser-2': 27}
    reprocess_production = await preparatory_report.get_reprocess_production_data(db, date_)
    for machine, production_dict in reprocess_production.items():
        column = reprocess_production_column_mappings.get(machine)
        if column:
            for row_index, single_date in enumerate(date_range, start=start_row):
                reprocess_data = production_dict.get(single_date.strftime('%Y-%m-%d'), 0)
                ws1.cell(row=row_index, column=column, value=reprocess_data)

    # ....................................... DEPTH ISSUE PRODUCTION

    #### ************************* MONTHLY PRODUCTION  REPORT DATA SHEET
    monthly_production = await preparatory_report.get_monthly_machines_production(db, date_)
    column_mappings = {'OSTHOFF-1': 2, 'OSTHOFF-2': 3, 'PTR': 4, 'NPR': 5, 'OMR': 6, 'NMR': 7, 'NPS': 8, 'Batcher-1': 9,
                       'Batcher-2': 10, 'OSTHOFF-1 BATCH': 11, 'X2': 12, 'X3': 13, 'LAFFER': 14, 'LAFFER-2': 15,
                       'LAFFER-3': 16, 'LAFFER-4': 17, 'LAFFER RASING': 18, 'LAFFER SHARING': 19, 'SOAPER-2': 20
                       }
    for machine, machine_data in monthly_production.items():
        c_idx = column_mappings.get(machine)
        if c_idx is not None:
            row_idx = 6
            for data_dict in machine_data:
                date_str = data_dict['date']
                formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
                ws2.cell(row=row_idx, column=1, value=formatted_date)
                ws2.cell(row=row_idx, column=c_idx, value=data_dict['production'])
                row_idx += 1

    ######### ******************* EFFECTIVE MONTHLY PRODUCTION

    effective_monthly_production = await preparatory_report.get_monthly_effective_production(db, date_)
    for machine, machine_data in effective_monthly_production.items():
        c_idx = kg_production_column_mappings.get(machine)
        if c_idx is not None:
            row_idx = 4
            for data_dict in machine_data:
                date_str = data_dict['date']
                formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
                ws3.cell(row=row_idx, column=1, value=formatted_date)
                ws3.cell(row=row_idx, column=c_idx, value=data_dict['production'])
                row_idx += 1

    ######## ******************* FRONT PAGE REPORT SHEET DATA
    ws4['A1'].value = f"PREP FRONT PAGE AS ON {date_}"
    ws4['A24'].value = f"PRODUCTION & UTILIZATION DETAIL AS ON {date_}"
    ws4['K46'].value = date_
    machine_data = await preparatory_report.get_daily_production_shift_wise(db, date_)
    fill_machine_data(26, ws4, machine_data)

    stopage_data = await crud.get_duration_per_machine(db, date_)
    column_mapping_stopage_data = {"MECH BRKD.": "B", "ELEC BRKD": "D", "CLG": "G", "CHANGE OVER": "H",
                                   "PM": "I", "NO STEAM OR NO AIR": "J", "POWER CUT": "K", "NO PROGRAM": "L",
                                   "NO MAN POWER": "M",
                                   "MISC": "N", "INSECT ISSUE": "P", }
    stopage_machine_data = stopage_data.get("machine_durations", {})
    await fill_stopage_data(ws4, stopage_machine_data, column_mapping_stopage_data)

    mtd_stopage_data = await crud.get_mtd_duration_machine_for_date_Range(db, date_)
    column_mapping_mtd_stopage_data = {"MTD  MECH  BRKD.": "C", "MTD  ELEC  BRKD.": "E"}
    mtd_machine_data = mtd_stopage_data.get("machine_durations", {})
    await fill_stopage_data(ws4, mtd_machine_data, column_mapping_mtd_stopage_data)

    #    ##### ******************* PR SHEET DATA
    #    # pr_production_data = await preparatory_report.get_pr_effective_production_data(db, date_)
    #    pr_production_data = await preparatory_report.get_effective_production_data(db, date_, ['Perble', 'PTR'])
    #    pr_machine_columns = {"Perble": (10, 11), "PTR": (18, 19)}
    #    await write_machine_data_to_worksheet(ws5, pr_production_data, pr_machine_columns, start_row)
    #
    #    # pr_reprocess_production_data = await preparatory_report.get_pr_reprocess_production_data(db, date_)
    #    pr_reprocess_production_data = await preparatory_report.get_reprocess_production_data_by_machine_list(db, date_,
    #                                                                                                          ['NPR',
    #                                                                                                           'PTR'])
    #    pr_reprocess_machine_columns = {"NPR": (14, 15), "PTR": (22, 23)}
    #    await write_machine_data_to_worksheet(ws5, pr_reprocess_production_data, pr_reprocess_machine_columns, start_row)

    ######## ******************* MR  OLD  SHEET DATA
    old_mr_production_data = await preparatory_report.get_effective_production_data(db, date_, ['Merceriser'])
    old_mr_machine_columns = {"Merceriser": (2, 3)}
    await write_machine_data_to_worksheet(ws7, old_mr_production_data, old_mr_machine_columns, 4)

    old_mr_reprocess_production_data = await preparatory_report.get_reprocess_production_data_by_machine_list(db, date_,
                                                                                                              ['NPR'])
    old_mr_reprocess_machine_columns = {"NPR": (7, 8)}
    await write_machine_data_to_worksheet(ws7, old_mr_reprocess_production_data, old_mr_reprocess_machine_columns, 4)

    ######## ******************* MR  NEW  SHEET DATA
    new_mr_production_data = await preparatory_report.get_effective_production_data(db, date_, ['Merceriser'])
    new_mr_machine_columns = {"Merceriser": (2, 3)}
    await write_machine_data_to_worksheet(ws8, new_mr_production_data, new_mr_machine_columns, 7)

    new_mr_reprocess_production_data = await preparatory_report.get_reprocess_production_data_by_machine_list(db, date_,
                                                                                                              [
                                                                                                                  'Merceriser'])
    new_mr_reprocess_machine_columns = {"Merceriser": (6, 7)}
    await write_machine_data_to_worksheet(ws8, new_mr_reprocess_production_data, new_mr_reprocess_machine_columns, 7)

    ######## ******************* NPS PRODUCTION  SHEET DATA
    nps_production_data = await preparatory_report.get_effective_production_data(db, date_, ['NPS'])
    nps_production_machine_columns = {"NPS": (2, 3)}
    await write_machine_data_to_worksheet(ws6, nps_production_data, nps_production_machine_columns, 18)

    path_ = os.path.join(dirname, f"Reports/Vardhman_Preparatory_Production_Report_{date_}.xlsx")
    wb.save(path_)

    return path_


async def get_preparatory_production_report_data(db: Session, date_: date):
    wb = load_workbook(os.path.join(dirname, f"Reports/Template/Preparatory_Production_Report_Template.xlsx"))
    ws1 = wb["MTD PROD KGS & RE MTRS_"]
    ws2 = wb["MTD PRODUTIN"]
    ws3 = wb["EFFECTIN PROB."]
    ws4 = wb["FRONT PAGE "]
    #    ws5 = wb["PR"]
    ##    ws6 = wb["NPS PROD"]
    #    ws7 = wb["MR OLD"]
    #    ws8 = wb["MR. NEW "]

    ##### ******************************* KG PRODUCTION REPORT SHEET
    production_data = await get_production_kg_of_each_machine(db, date_)
    start_row = 5
    date_range = pd.date_range(datetime(date_.year, date_.month, 1),
                               datetime(date_.year, date_.month, calendar.monthrange(date_.year, date_.month)[1]))

    for row_index, single_date in enumerate(date_range, start=start_row):
        ws1[f"A{row_index}"] = single_date.strftime('%d-%b-%y')
        ws1[f"W{row_index}"] = single_date.strftime('%d-%b-%y')

    kg_production_column_mappings = {'SANDO': 2, 'Osthoff-1': 3, 'Osthoff-2': 4, 'PTR': 5, 'Perble': 6,
                                     'Merceriser-1': 7, 'Merceriser-2': 8,
                                     'NPS': 9, 'Batcher-1': 10, 'Batcher-2': 11, 'Osthoff-1 BATCH': 12, 'Xetma-2': 13,
                                     'Xetma-1': 14,
                                     'Lafer-1': 15, 'Lafer-2': 16, 'Lafer-3': 17, 'Lafer-4': 18,
                                     'Lafer-5-Raising': 19,
                                     'Lafer-5-Shearing': 20, 'Soaper': 21
                                     }
    for machine_data in production_data:

        for machine, production_dict in machine_data.items():
            column = kg_production_column_mappings.get(machine)

            if column:
                for row_index, single_date in enumerate(date_range, start=start_row):
                    # production_kg = production_dict.get(single_date.strftime('%Y-%m-%d'), 0)
                    production_kg = production_dict.get(single_date.strftime('%d-%b-%Y'), 0)
                    ws1.cell(row=row_index, column=column, value=production_kg)

    # ....................................... REPROCESS MTR PRODUCTION
    print("hsa")
    reprocess_production_column_mappings = {'PTR': 24, 'Perble': 25, 'Merceriser-1': 26, 'Merceriser-2': 27}
    reprocess_production = await preparatory_report.get_reprocess_production_data(db, date_)
    for machine, production_dict in reprocess_production.items():
        column = reprocess_production_column_mappings.get(machine)
        if column:
            for row_index, single_date in enumerate(date_range, start=start_row):
                # reprocess_data = production_dict.get(single_date.strftime('%Y-%m-%d'), 0)
                reprocess_data = production_dict.get(single_date.strftime('%d-%b-%Y'), 0)
                ws1.cell(row=row_index, column=column, value=reprocess_data)

    # ....................................... DEPTH ISSUE PRODUCTION
    print("chd")
    ### ************************* MONTHLY PRODUCTION  REPORT DATA SHEET
    monthly_production = await preparatory_report.get_monthly_machines_production(db, date_)
    column_mappings = {'Osthoff-1': 2, 'Osthoff-2': 3, 'PTR': 4, 'Perble': 5, 'Merceriser-1': 6, 'Merceriser-2': 7,
                       'NPS': 8, 'Batcher-1': 9,
                       'Batcher-2': 10, 'Osthoff-1 BATCH': 11, 'Xetma-2': 12, 'Xetma-1': 13, 'Lafer-1': 14,
                       'Lafer-2': 15,
                       'Lafer-3': 16, 'Lafer-4': 17, 'Lafer-5-Raising': 18, 'Lafer-5-Shearing': 19, 'Soaper': 20
                       }
    for machine, machine_data in monthly_production.items():
        c_idx = column_mappings.get(machine)
        if c_idx is not None:
            row_idx = 2
            for data_dict in machine_data:
                date_str = data_dict['date']
                formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
                ws2.cell(row=row_idx, column=1, value=formatted_date)
                ws2.cell(row=row_idx, column=c_idx, value=data_dict['production'])
                row_idx += 1

    ######### ******************* EFFECTIVE MONTHLY PRODUCTION
    print("CHD")
    effective_monthly_production = await preparatory_report.get_monthly_effective_production(db, date_)
    for machine, machine_data in effective_monthly_production.items():
        c_idx = kg_production_column_mappings.get(machine)
        if c_idx is not None:
            row_idx = 4
            for data_dict in machine_data:
                date_str = data_dict['date']
                formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d-%b-%y')
                ws3.cell(row=row_idx, column=1, value=formatted_date)
                ws3.cell(row=row_idx, column=c_idx, value=data_dict['production'])
                row_idx += 1

    ######## ******************* FRONT PAGE REPORT SHEET DATA
    ws4['A1'].value = f"PREP FRONT PAGE AS ON {date_}"
    ws4['A24'].value = f"PRODUCTION & UTILIZATION DETAIL AS ON {date_}"
    ws4['K46'].value = date_
    machine_data = await preparatory_report.get_daily_production_shift_wise(db, date_)
    fill_machine_data(26, ws4, machine_data)

    stopage_data = await crud.get_duration_per_machine(db, date_)
    column_mapping_stopage_data = {"MECH BRKD.": "B", "ELEC BRKD": "D", "CLG": "G", "CHANGE OVER": "H",
                                   "PM": "I", "NO STEAM OR NO AIR": "J", "POWER CUT": "K", "NO PROGRAM": "L",
                                   "NO MAN POWER": "M",
                                   "MISC": "N", "INSECT ISSUE": "P", }
    stopage_machine_data = stopage_data.get("machine_durations", {})
    await fill_stopage_data(ws4, stopage_machine_data, column_mapping_stopage_data)

    mtd_stopage_data = await crud.get_mtd_duration_machine_for_date_Range(db, date_)
    column_mapping_mtd_stopage_data = {"MTD  MECH  BRKD.": "C", "MTD  ELEC  BRKD.": "E"}
    mtd_machine_data = mtd_stopage_data.get("machine_durations", {})
    await fill_stopage_data(ws4, mtd_machine_data, column_mapping_mtd_stopage_data)

    #    ##### ******************* PR SHEET DATA
    #    # pr_production_data = await preparatory_report.get_pr_effective_production_data(db, date_)
    #    pr_production_data = await preparatory_report.get_effective_production_data(db, date_, ['Perble', 'PTR'])
    #    pr_machine_columns = {"Perble": (10, 11), "PTR": (18, 19)}
    #    await write_machine_data_to_worksheet(ws5, pr_production_data, pr_machine_columns, start_row)
    #
    #    # pr_reprocess_production_data = await preparatory_report.get_pr_reprocess_production_data(db, date_)
    #    pr_reprocess_production_data = await preparatory_report.get_reprocess_production_data_by_machine_list(db, date_,
    #                                                                                                          ['Perble',
    #                                                                                                           'PTR'])
    #    pr_reprocess_machine_columns = {"Perble": (14, 15), "PTR": (22, 23)}
    #    await write_machine_data_to_worksheet(ws5, pr_reprocess_production_data, pr_reprocess_machine_columns, start_row)

    #    ######## ******************* MR  OLD  SHEET DATA
    #    old_mr_production_data = await preparatory_report.get_effective_production_data(db, date_, ['Merceriser'])
    #    old_mr_machine_columns = {"Merceriser": (2, 3)}
    #    await write_machine_data_to_worksheet(ws7, old_mr_production_data, old_mr_machine_columns, 4)
    #
    #    old_mr_reprocess_production_data = await preparatory_report.get_reprocess_production_data_by_machine_list(db, date_,
    #                                                                                                              ['Perble'])
    #    old_mr_reprocess_machine_columns = {"Perble": (7, 8)}
    #    await write_machine_data_to_worksheet(ws7, old_mr_reprocess_production_data, old_mr_reprocess_machine_columns, 4)
    #
    #    ######## ******************* MR  NEW  SHEET DATA
    #    new_mr_production_data = await preparatory_report.get_effective_production_data(db, date_, ['Merceriser'])
    #    new_mr_machine_columns = {"Merceriser": (2, 3)}
    #    await write_machine_data_to_worksheet(ws8, new_mr_production_data, new_mr_machine_columns, 7)
    #
    #    new_mr_reprocess_production_data = await preparatory_report.get_reprocess_production_data_by_machine_list(db, date_,
    #                                                                                                              [
    #                                                                                                                  'Merceriser'])
    #    new_mr_reprocess_machine_columns = {"Merceriser": (6, 7)}
    #    await write_machine_data_to_worksheet(ws8, new_mr_reprocess_production_data, new_mr_reprocess_machine_columns, 7)

    #    ######## ******************* NPS PRODUCTION  SHEET DATA
    #    nps_production_data = await preparatory_report.get_effective_production_data(db, date_, ['NPS'])
    #    nps_production_machine_columns = {"NPS": (2, 3)}
    #    await write_machine_data_to_worksheet(ws6, nps_production_data, nps_production_machine_columns, 18)

    path_ = os.path.join(dirname, f"Reports/Vardhman_Preparatory_Production_Report_{date_}.xlsx")
    wb.save(path_)

    return path_


async def generate_preparatory_report_data(db: Session, date_: date):
    wb = load_workbook(os.path.join(dirname, "Reports/Template/Preparatory_Report_Template.xlsx"))
    current_month = datetime.now().strftime('%B')
    current_year = datetime.now().year

    if "FRONT PAGE" in wb.sheetnames:
        front_page = wb["FRONT PAGE"]
        heading_text = f"PERFORMANCE REVIEW - PREPARATORY DEPARTMENT {current_month} {current_year}"
        font_style = Font(size=18, bold=True)
        alignment_center = Alignment(horizontal="center", vertical="center")

        for cell in ['A1', 'Q1', 'AM1']:
            front_page[cell].value = heading_text
            front_page[cell].font = font_style
            front_page[cell].alignment = alignment_center

    # List of machines
    machines_temp2 = ['Osthoff-1', 'Osthoff-2', 'PTR', 'NPS', 'Batcher-1', 'Batcher-2', 'Perble',
                      'Lafer-1', 'Lafer-2', 'Lafer-3', 'Lafer-4', 'Lafer-5-Raising',
                      'Lafer-5-Shearing', 'Soaper', 'Swastik', 'Merceriser-1', 'Merceriser-2', 'Xetma-1', 'Xetma-2']

    col_mappings = {
        'process_demand': {'Process Demand': 3},
        'rep_prod': {'Reprocess': 4}
    }
    col_dur = {
        'mech_dur': {'Mechanical Breakdown': 11},
        'elec_dur': {'Electrical Breakdown': 12},
        'mach_clean': {'Machine cleaning': 13},
        'changeover': {'Fabric changeover': 14},
        'pm': {'Preventive Maintenance': 15},
        'no_power': {'No Power': 16},
        'no_steam': {'No Steam': 17},
        'no_program': {'No Program': 18},
        'man_power_shortage': {'Man power shortage': 19},
        'insect_problem': {'Insect Problem': 21},
        'no_trolley': {'No Trolley': 22}
    }

    for machine in machines_temp2:
        # Handle special case for Xetma-1 mapping to Xetma-3
        sheet_name = 'Xetma-3' if machine == 'Xetma-1' else machine

        # Check if the dynamically generated sheet name exists
        if sheet_name not in wb.sheetnames:
            print(f"Sheet for {sheet_name} not found.")
            continue

        # Access the correct sheet name
        new_sheet = wb[sheet_name]
        new_sheet['A1'].value = f'{machine} Production'
        new_sheet['J1'].value = f"{date_.strftime('%B')} {date_.year}"

        # Machine-specific data
        machine_data_production = [(machine, "B")]
        machine_data_duration = [(machine, "I")]
        production_till_date = [(machine, "F")]
        dur_till_date = [(machine, "J")]
        production_in_kg = [(machine, "G")]
        cumulative_production = [(machine, "H")]

        await populate_machine_data(new_sheet, 4, machine_data_production,
                                    preparatory_report.fetch_machine_production_for_month, db, date_)

        await populate_machine_data(new_sheet, 4, production_in_kg,
                                    preparatory_report.fetch_prod_kgs_machine_production_for_month, db, date_)

        await populate_machine_data(new_sheet, 4, machine_data_duration,
                                    preparatory_report.fetch_machine_duration_for_month, db, date_)

        await populate_machine_data(new_sheet, 4, production_till_date,
                                    preparatory_report.fetch_machine_production_till_date, db, date_)

        await populate_machine_data(new_sheet, 4, dur_till_date, preparatory_report.fetch_machine_duration_till_date,
                                    db, date_)

        ##.......................

        stop_list = ['Machanical Breakdown', 'Electrical Breakdown', 'Machine cleaning', 'Fabric changeover',
                     'Preventive Maintenance', 'No Power', 'No Steam', 'Man Power Shortage', 'No Program',
                     'Insect Problem', 'No Trolley']

        month_production = await preparatory_report.fetch_misc_duration_daily(db, date_, machine, stop_list)

        start_row = 4
        column = 20

        for i, (date_str, value) in enumerate(month_production.items(), start=start_row):
            new_sheet.cell(row=i, column=column).value = value

        rep_prod = await preparatory_report.fetch_operation_production_for_month(db, date_, machine, 'Reprocess')
        populate_cells(new_sheet, rep_prod, 4, col_mappings['rep_prod'])

        process_demand = await preparatory_report.fetch_operation_production_for_month(db, date_, machine,
                                                                                       'Process Demand')
        populate_cells(new_sheet, process_demand, 4, col_mappings['process_demand'])

        mech_dur = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                               'Machanical Breakdown')
        populate_cells(new_sheet, mech_dur, 4, col_dur['mech_dur'])

        elec_dur = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                               'Electrical Breakdown')
        populate_cells(new_sheet, elec_dur, 4, col_dur['elec_dur'])

        mach_clean = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'Machine cleaning')
        populate_cells(new_sheet, mach_clean, 4, col_dur['mach_clean'])

        changeover = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                                 'Fabric changeover')
        populate_cells(new_sheet, changeover, 4, col_dur['changeover'])

        pm = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'Preventive Maintenance')
        populate_cells(new_sheet, pm, 4, col_dur['pm'])

        no_power = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Power')
        populate_cells(new_sheet, no_power, 4, col_dur['no_power'])

        no_steam = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Steam')
        populate_cells(new_sheet, no_steam, 4, col_dur['no_steam'])

        no_program = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Program')
        populate_cells(new_sheet, no_program, 4, col_dur['no_program'])

        man_power_shortage = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                                         'Man Power Shortage')
        populate_cells(new_sheet, man_power_shortage, 4, col_dur['man_power_shortage'])

        insect_problem = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine,
                                                                                     'Insect Problem')
        populate_cells(new_sheet, insect_problem, 4, col_dur['insect_problem'])

        no_trolley = await preparatory_report.fetch_operation_duration_for_month(db, date_, machine, 'No Trolley')
        populate_cells(new_sheet, no_trolley, 4, col_dur['no_trolley'])

    path_ = os.path.join(dirname, f"Reports/Vardhman_Preparartory_Report_{date_}.xlsx")
    wb.save(path_)
    print("Report saved.")

    return path_
