import os
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi.responses import FileResponse
from fastapi import Depends, FastAPI, HTTPException, UploadFile
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import cast, Time, or_, and_, func
from datetime import date, datetime
from sqlalchemy import select
from .analytics import calculate_key_value, get_po_details, get_po_data
from .. import crud, models, schemas
from ..database import SessionLocal, engine
import pytz
from fastapi import Depends, APIRouter
import logging
from collections import defaultdict
from io import BytesIO
from fastapi.responses import StreamingResponse

log = logging.getLogger("uvicorn")
log.setLevel(logging.INFO)

IST = pytz.timezone('Asia/Kolkata')

if getattr(sys, 'frozen', False):
    dirname = os.path.dirname(sys.executable)
else:
    dirname = os.path.dirname(os.path.abspath(__file__))


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


router = APIRouter(tags=["Report"])


async def report(from_date: date, to_date: date, db: Session = Depends(get_db)):
    po_data = db.query(models.PoData).filter(models.PoData.date_.between(from_date, to_date),
                                             models.PoData.stop_time != None).order_by(models.PoData.id.asc()).all()

    report_data = []

    for item in po_data:
        keys = await calculate_key_value(db=db, po_uuid=item.po_uuid)
        #Total Breakdown Duration
        total_breakdown_duration = db.query(func.coalesce(func.sum(models.BreakdownData.duration), 0)).filter(
            models.BreakdownData.breakdown_po_uuid == item.po_uuid,
            models.BreakdownData.machine_name == item.machine_name,
            models.BreakdownData.line == item.line).scalar()

        report_data.append({**item.__dict__, **keys, "breakdown_duration": total_breakdown_duration})
    return report_data


@router.get("/report/generate_history_po_data_report/{from_date}/{to_date}")
async def generate_history_po_data_report(from_date: date, to_date: date, db: Session = Depends(get_db)):
    try:
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(base_dir, "Reports", "Template", "history_po_data_template.xlsx")
        wb = load_workbook(template_path)
        ws = wb.active
        data = await report(from_date, to_date, db)

        for item in data:
            item.pop("id", None)

        headers = ["date_", "machine_name", "line", "shift", "po_number", "section", "category",
                   "machine_speed", "machine_speed_unit", "operation", "start_time", "stop_time", "duration",
                   "operator_name", "target_length", "target_unit"]
        end_header = ["is_partial_gr", "is_complete"]
        dynamic_headers = set()

        for item in data:
            for key in item.keys():
                if key.startswith("last_"):
                    dynamic_headers.add(key)

        headers = headers + sorted(dynamic_headers) + end_header

        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header.upper()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        for row_num, item in enumerate(data, start=3):
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = item.get(header)
                cell.alignment = center_alignment
                cell.border = thin_border

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                cell = row[0]
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 5

        file_path = os.path.join(base_dir, "Reports", "history_po_data_report.xlsx")
        wb.save(file_path)
        return FileResponse(path=file_path,
                            filename=f"PO Data Report_{from_date}_to_{to_date}.xlsx",
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return {"error": str(e)}


# @router.get("/report/get_daily_production_data")
async def get_daily_production_data(from_date: date, to_date: date, db: Session = Depends(get_db)):
    machine_list = db.scalars(select(models.PoData.machine_name).distinct()).all()
    all_key = db.scalars(select(models.HourlyData.key).distinct()).all()
    po_data = await get_po_data(from_date=from_date, to_date=to_date, db=db)
    if not po_data:
        return {"from_date": from_date,
                "to_date": to_date,
                "avaliable keys": all_key,
                "machine": machine_list,
                "data": []}

    result = []
    for data in po_data:
        keys = await get_po_details(po_uuid=data["po_uuid"], db=db)
        result.append(keys)
    return {"from_date": from_date,
            "to_date": to_date,
            "avaliable keys": all_key,
            "machine": machine_list,
            "data": result}


def generate_day_report(response_data, output_file="Day_Report.xlsx"):
    """
    response_data = API Response
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    from_date = response_data["from_date"]
    to_date = response_data["to_date"]

    machine_data = defaultdict(list)

    # Group machine wise
    for item in response_data["data"]:
        machine_data[item["machine_name"]].append(item)

    thin = Side(border_style="thin", color="000000")

    header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    summary_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    for machine, records in machine_data.items():
        ws = wb.create_sheet(title=machine[:31])

        ###################################################
        # Collect Dynamic Keys
        ###################################################

        dynamic_keys = set()

        for rec in records:
            dynamic_keys.update(rec.get("key", {}).keys())

        # average_speed already available separately
        dynamic_keys.discard("average_speed")
        dynamic_keys = sorted(dynamic_keys)

        ###################################################
        # Fixed Header
        ###################################################

        headers = ["DATE", "SHIFT", "PO NUMBER", "START TIME", "STOP TIME", "DURATION (SEC)", "CATEGORY", "OPERATION",
                   "SECTION", "LINE", "OPERATOR NAME", "TARGET LENGTH", "TARGET UNIT", "MACHINE SPEED",
                   "AVERAGE SPEED", ]
        headers.extend([k.upper().replace("_", " ") for k in dynamic_keys])
        total_cols = len(headers)

        ###################################################
        # Title
        ###################################################

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        cell = ws.cell(1, 1)
        cell.value = f"DAY REPORT FROM {from_date} TO {to_date}"
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = title_fill

        ###################################################
        # Summary
        ###################################################

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
        cell = ws.cell(3, 1)
        cell.value = f"{machine} BATCH SUMMARY"
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = summary_fill

        ###################################################
        # Header Row
        ###################################################

        header_row = 5

        for col, header in enumerate(headers, start=1):
            c = ws.cell(header_row, col)
            c.value = header.upper()
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ###################################################
        # Data
        ###################################################
        row_no = 6

        def format_datetime(value):
            if not value:
                return ""
            try:
                return datetime.fromisoformat(value).strftime("%d-%m-%Y %H:%M:%S")
            except:
                return value

        for item in records:
            row = [
                item.get("date_", ""),
                item.get("shift", ""),
                item.get("po_number", ""),
                format_datetime(item.get("start_time")),
                format_datetime(item.get("stop_time")),
                item.get("duration", ""),
                item.get("category", ""),
                item.get("operation", ""),
                item.get("section", ""),
                item.get("line", ""),
                item.get("operator_name", ""),
                item.get("target_length", ""),
                item.get("target_unit", ""),
                item.get("machine_speed", ""),
                item.get("key", {}).get("average_speed", "")
            ]

            # Dynamic key values
            for key in dynamic_keys:
                row.append(item.get("key", {}).get(key, ""))

            for col, value in enumerate(row, start=1):
                cell = ws.cell(row_no, col)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            row_no += 1

        ws.freeze_panes = "A6"

        ###################################################
        # Auto Width
        ###################################################

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/report/download_daywise_report")
async def download_day_report(from_date: date, to_date: date, db: Session = Depends(get_db)):
    response_data = await get_daily_production_data(from_date=from_date, to_date=to_date, db=db)
    # CRITICAL FIX: Raise an exception if no data exists to prevent Excel save crash
    if not response_data.get("data"):
        raise HTTPException(
            status_code=404,
            detail=f"No production data found between {from_date} and {to_date} to generate a report."
        )
    excel = generate_day_report(response_data)
    return StreamingResponse(excel,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition":
                                          f'attachment; filename="Day_Report_{from_date}_{to_date}.xlsx"'
                                      }
                             )


AVERAGE_KEYS = {"Speed", "average_speed"}


def calculate_key_value_for_report(rows):
    grouped = defaultdict(list)

    # Group rows by key
    for row in rows:
        grouped[row.key].append(row)

    result = {}

    # Calculate values for every key
    for key, values in grouped.items():

        first = values[0].key_start
        last = values[-1].key_stop

        if key in {"Length", "Speed"}:
            result[key] = last or 0

        else:
            try:
                result[key] = round((last or 0) - (first or 0), 2)
            except Exception:
                result[key] = 0

    # Default average speed
    result["average_speed"] = 0

    # Calculate Average Speed
    if "Length" in grouped:

        values = grouped["Length"]

        first = values[0]
        last = values[-1]

        if (
                first.created_at
                and last.updated_at
                and last.updated_at > first.created_at
        ):
            duration = (
                               last.updated_at - first.created_at
                       ).total_seconds() / 60

            total_length = last.key_stop or 0

            result["average_speed"] = (
                round(total_length / duration, 2)
                if duration > 0 else 0
            )

    return result


@router.get("/report/generate_summary_monthly_data")
async def generate_summary_monthly_report(from_date: date, to_date: date, db: Session = Depends(get_db)):
    machine_list = sorted(
        [machine for machine in db.scalars(select(models.PoData.machine_name).distinct()).all() if machine])

    key_list = sorted([key for key in db.scalars(select(models.HourlyData.key).distinct()).all() if key])

    if "average_speed" not in key_list:
        key_list.append("average_speed")

    # -----------------------------
    # Fetch All PO Data
    # -----------------------------
    po_data = db.query(models.PoData).filter(models.PoData.date_.between(from_date, to_date),
                                             models.PoData.stop_time.isnot(None)).order_by(models.PoData.date_,
                                                                                           models.PoData.machine_name,
                                                                                           models.PoData.id).all()

    if not po_data:
        return {
            "from_date": from_date,
            "to_date": to_date,
            "machine_list": machine_list,
            "key_list": key_list,
            "datewise": [],
            "overall_summary": []
        }

    # -----------------------------
    # Fetch All HourlyData (Single Query)
    # -----------------------------
    po_uuid_list = [po.po_uuid for po in po_data]

    hourly_data = db.query(models.HourlyData).filter(models.HourlyData.po_uuid.in_(po_uuid_list)).order_by(
        models.HourlyData.po_uuid, models.HourlyData.key, models.HourlyData.id).all()

    hourly_by_po = defaultdict(list)

    for row in hourly_data:
        hourly_by_po[row.po_uuid].append(row)

    # -----------------------------
    # Containers
    # -----------------------------
    datewise = {}
    overall_summary = {}

    date_avg_values = defaultdict(lambda: defaultdict(list))
    machine_avg_values = defaultdict(lambda: defaultdict(list))

    # -----------------------------
    # Process Every PO
    # -----------------------------
    for po in po_data:
        values = calculate_key_value_for_report(hourly_by_po.get(po.po_uuid, []))
        current_date = po.start_time.date()
        machine = po.machine_name

        # -----------------------------
        # Date Initialization
        # -----------------------------
        if current_date not in datewise:
            datewise[current_date] = {"date": current_date, "po_count": 0}
        datewise[current_date]["po_count"] += 1

        # -----------------------------
        # Machine Initialization
        # -----------------------------
        if machine not in overall_summary:
            overall_summary[machine] = {"machine_name": machine, "po_count": 0}

        overall_summary[machine]["po_count"] += 1
        duration = round(float(po.duration or 0), 2)

        datewise[current_date]["machine_running"] = round(datewise[current_date].get("machine_running", 0) + duration,
                                                          2)
        overall_summary[machine]["machine_running"] = round(
            overall_summary[machine].get("machine_running", 0) + duration, 2)
        # -----------------------------
        # Update Summary
        # -----------------------------
        for key, value in values.items():
            if value is None:
                continue

            if key in AVERAGE_KEYS:
                date_avg_values[current_date][key].append(value)
                machine_avg_values[machine][key].append(value)
            else:
                datewise[current_date][key] = round((datewise[current_date].get(key, 0) + value), 2)
                overall_summary[machine][key] = round((overall_summary[machine].get(key, 0) + value), 2)
    # -----------------------------------------
    # Calculate Date-wise Average Values
    # -----------------------------------------
    for current_date, keys in date_avg_values.items():
        for key, values in keys.items():
            if values:
                datewise[current_date][key] = round(sum(values) / len(values), 2)
            else:
                datewise[current_date][key] = 0

    # -----------------------------------------
    # Calculate Machine-wise Average Values
    # -----------------------------------------
    for machine, keys in machine_avg_values.items():
        for key, values in keys.items():
            if values:
                overall_summary[machine][key] = round(sum(values) / len(values), 2)
            else:
                overall_summary[machine][key] = 0

    # -----------------------------------------
    # Ensure every key exists in Datewise
    # -----------------------------------------
    for item in datewise.values():
        for key in key_list:
            if key not in item:
                item[key] = 0

    # -----------------------------------------
    # Ensure every key exists in Machine Summary
    # -----------------------------------------
    for item in overall_summary.values():
        for key in key_list:
            if key not in item:
                item[key] = 0

    datewise_result = sorted(datewise.values(), key=lambda x: x["date"])
    overall_result = sorted(overall_summary.values(), key=lambda x: x["machine_name"])

    return {
        "from_date": from_date,
        "to_date": to_date,
        "machine_list": machine_list,
        "key_list": key_list,
        "datewise": datewise_result,
        "overall_summary": overall_result,
    }


def generate_month_report(response_data):
    """
    Generate Excel report in memory and return a StreamingResponse.

    NOTE ON write_only=True:
    WriteOnlyWorksheet (used when Workbook(write_only=True)) does not
    implement merge_cells, column_dimensions, or auto_filter at all
    verified directly against the installed openpyxl version. Only
    freeze_panes survives. Since merged titles, auto-width columns, and
    auto-filter are hard requirements here, this function uses a standard
    Workbook(). For the row counts involved (a handful of machines and a
    month's worth of datewise rows each), this has no meaningful
    performance cost.
    """

    wb = Workbook()
    wb.remove(wb.active)

    from_date = response_data["from_date"]
    to_date = response_data["to_date"]

    overall_summary = response_data["overall_summary"]
    datewise = response_data["datewise"]

    # -----------------------------------
    # Styles
    # -----------------------------------

    title_font = Font(size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")

    heading_font = Font(size=15, bold=True, color="000000")
    heading_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    center = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # -----------------------------------
    # Helpers
    # -----------------------------------

    def write_merged_banner(ws, text, row, total_cols, font, fill):
        """Write a single styled value merged across total_cols on `row`."""
        last_col = get_column_letter(total_cols)
        ws.merge_cells(f"A{row}:{last_col}{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        # paint the whole merged range so it looks solid, not just col A
        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).fill = fill
        return row + 1

    def write_header_row(ws, headers, row):
        """Write an uppercase, bold, bordered, filled header row. Returns next row."""
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=key.upper().replace("_", " "))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        return row + 1

    def write_data_rows(ws, headers, rows, row):
        """Write plain bordered data rows for the given list of dicts. Returns next row."""
        for record in rows:
            for col_idx, key in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=record.get(key, ""))
                cell.alignment = center
                cell.border = border
            row += 1
        return row

    def autofit_columns(ws):
        """Auto-width every column based on its longest rendered value."""
        for column_cells in ws.columns:
            length = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in column_cells
            )
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

    # ==========================================
    # OVERALL SUMMARY SHEET
    # ==========================================

    ws = wb.create_sheet("OVERALL SUMMARY")

    overall_headers = list(overall_summary[0].keys())
    total_columns = len(overall_headers)
    last_col = get_column_letter(total_columns)

    # Row 1: Title (merged)
    row = write_merged_banner(
        ws, f"MONTH TILL DATE REPORT FROM {from_date} TO {to_date}",
        row=1, total_cols=total_columns, font=title_font, fill=title_fill
    )

    # blank row
    row += 1

    # Row 3: OVERALL SUMMARY heading (merged)
    row = write_merged_banner(
        ws, "OVERALL SUMMARY", row=row, total_cols=total_columns,
        font=heading_font, fill=heading_fill
    )

    # Row 4: headers
    header_row = row
    row = write_header_row(ws, overall_headers, row)

    # Row 5+: data
    row = write_data_rows(ws, overall_headers, overall_summary, row)
    last_row = row - 1

    ws.freeze_panes = "A6"
    autofit_columns(ws)

    # ==========================================
    # MACHINE SHEETS
    # ==========================================

    machine_datewise = defaultdict(list)

    for record in datewise:
        # if API response contains machine_name, use it
        machine = record.get("machine_name")

        # fallback for single-machine payloads (like the sample response)
        if machine is None:
            machine = overall_summary[0]["machine_name"]

        machine_datewise[machine].append(record)

    for machine_summary in overall_summary:

        machine = machine_summary["machine_name"]
        ws = wb.create_sheet(machine[:31])

        summary_headers = list(machine_summary.keys())
        datewise_rows = machine_datewise.get(machine, [])
        datewise_headers = list(datewise_rows[0].keys()) if datewise_rows else []

        total_columns = max(len(summary_headers), len(datewise_headers))
        last_col = get_column_letter(total_columns)

        # Row 1: Title (merged)
        row = write_merged_banner(
            ws, f"MONTH TILL DATE REPORT FROM {from_date} TO {to_date}",
            row=1, total_cols=total_columns, font=title_font, fill=title_fill
        )

        # blank row
        row += 1

        # Row 3: Machine name heading (merged)
        row = write_merged_banner(
            ws, machine.upper(), row=row, total_cols=total_columns,
            font=heading_font, fill=heading_fill
        )

        # Row 4: machine summary headers

        # DATEWISE SUMMARY heading (merged)
        row = write_merged_banner(
            ws, "DATEWISE SUMMARY", row=row, total_cols=total_columns,
            font=heading_font, fill=heading_fill
        )

        if datewise_headers:
            datewise_header_row = row
            row = write_header_row(ws, datewise_headers, row)
            row = write_data_rows(ws, datewise_headers, datewise_rows, row)
            datewise_last_row = row - 1

        ws.freeze_panes = "A6"
        autofit_columns(ws)

    # ==========================================
    # Return Bytes
    # ==========================================

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Month_Report.xlsx"'
        },
    )


@router.get("/report/generate_summary_monthly_report")
async def download_monthly_report(from_date: date,to_date: date,db: Session = Depends(get_db)):
    response_data = await generate_summary_monthly_report(from_date, to_date, db)
    return generate_month_report(response_data)
